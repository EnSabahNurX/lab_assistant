import os
import glob
import shutil
import json
import tkinter as tk
import numpy as np
import traceback
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from tkinter import (
    ttk,
    messagebox,
)
from openpyxl import load_workbook
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl.worksheet.protection import SheetProtection


class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        self.widget.bind("<Enter>", self.show_tip)
        self.widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event=None):
        if self.tip_window or not self.text:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            tw,
            text=self.text,
            justify=tk.LEFT,
            background="#ffffe0",
            relief=tk.SOLID,
            borderwidth=1,
            font=("Helvetica", 10),
        )
        label.pack()

    def hide_tip(self, event=None):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None


class ExcelToJsonConverter:
    def __init__(self):
        # Constantes para process_grafik
        self.MIN_LIMIT_ROW = 51
        self.MAX_LIMIT_ROW = 55
        self.PRESSURE_DATA_START_ROW = 60
        self.MIN_COLUMN = 3
        self.MAX_COLUMN = 151

        self.root = tk.Tk()
        self.root.title("Ballistic Tests Converter")
        self.root.geometry("1200x700")
        self.root.minsize(1000, 600)
        self.root.columnconfigure(0, weight=1)
        self.root.columnconfigure(1, weight=3)
        self.root.rowconfigure(0, weight=1)

        # Left Frame (Database)
        self.frame_db = tk.Frame(self.root, relief="groove", padx=10, pady=10, bd=2)
        self.frame_db.grid(row=0, column=0, sticky="nsew")
        self.frame_db.columnconfigure(0, weight=1)
        self.frame_db.rowconfigure(5, weight=1)

        # Database Title Frame
        title_frame = tk.Frame(self.frame_db)
        title_frame.grid(row=0, column=0, sticky="ew")
        title_frame.columnconfigure(0, weight=1)

        self.label_database_title = tk.Label(
            title_frame, text="Database", font=("Helvetica", 14, "bold")
        )
        self.label_database_title.grid(row=0, column=0, sticky="w", pady=(0, 10))

        # Export Database Button
        self.btn_export_db = tk.Button(
            title_frame,
            text="Export Database",
            command=self.export_database_to_excel,
            font=("Helvetica", 12, "bold"),
            bg="#90EE90",
            fg="white",
            padx=15,
            pady=10,
        )
        self.btn_export_db.grid(row=0, column=1, sticky="e", pady=(0, 10))
        ToolTip(self.btn_export_db, "Export database to Excel")

        # Orders Input
        tk.Label(self.frame_db, text="Enter order numbers separated by commas:").grid(
            row=1, column=0, sticky="w"
        )
        self.entry_orders = tk.Entry(self.frame_db)
        self.entry_orders.grid(row=2, column=0, sticky="ew", pady=5)

        # Process and Remove Orders Buttons Side by Side
        btns_frame = tk.Frame(self.frame_db)
        btns_frame.grid(row=3, column=0, sticky="ew", pady=5)
        btns_frame.columnconfigure((0, 1), weight=1)

        self.btn_process = tk.Button(
            btns_frame, text="Add Entered Orders", command=self.process_orders
        )
        self.btn_process.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        ToolTip(self.btn_process, "Process and add orders to the database")

        self.btn_remove_orders = tk.Button(
            btns_frame,
            text="Remove Entered Orders",
            command=self.remove_orders_by_input,
        )
        self.btn_remove_orders.grid(row=0, column=1, sticky="ew", padx=(5, 0))
        ToolTip(self.btn_remove_orders, "Remove specified orders from the database")

        self.status_label = tk.Label(self.frame_db, text="", anchor="w", fg="green")
        self.status_label.grid(row=4, column=0, sticky="ew", pady=(0, 10))

        # Orders Manager
        self.frame_orders_manager = tk.Frame(
            self.frame_db, relief="groove", bd=2, padx=10, pady=10
        )
        self.frame_orders_manager.grid(row=5, column=0, sticky="nsew")
        self.frame_orders_manager.columnconfigure(0, weight=1)
        self.frame_orders_manager.rowconfigure(2, weight=1)

        tk.Label(self.frame_orders_manager, text="Orders:").grid(
            row=0, column=0, sticky="w", padx=5, pady=5
        )

        # Pagination and Filters
        self.current_page = 1
        self.orders_per_page = 10
        self.total_pages = 1

        self.pagination_frame = tk.Frame(self.frame_orders_manager)
        self.pagination_frame.grid(row=1, column=0, sticky="ew", padx=5)
        self.pagination_frame.columnconfigure((0, 1, 2, 3, 4, 5, 6), weight=1)

        tk.Label(self.pagination_frame, text="Items per page:").grid(
            row=0, column=0, sticky="w"
        )
        self.page_selector = ttk.Combobox(
            self.pagination_frame,
            values=[5, 10, 15, 20, 25, 30],
            width=5,
            state="readonly",
        )
        self.page_selector.set(self.orders_per_page)
        self.page_selector.grid(row=0, column=1, sticky="w")
        self.page_selector.bind("<<ComboboxSelected>>", self.on_items_per_page_changed)

        self.select_all_var = tk.BooleanVar()
        self.select_all_chk = tk.Checkbutton(
            self.pagination_frame,
            text="Select All",
            variable=self.select_all_var,
            command=self.toggle_select_all,
        )
        self.select_all_chk.grid(row=0, column=2, sticky="w", padx=10)

        self.version_var = tk.StringVar()
        self.version_combobox = ttk.Combobox(
            self.pagination_frame,
            textvariable=self.version_var,
            state="readonly",
            width=12,
        )
        self.version_combobox.grid(row=0, column=3, sticky="w", padx=10)
        self.version_combobox.bind("<<ComboboxSelected>>", self.on_version_filter)
        self.version_combobox["values"] = []
        self.version_combobox.set("All")

        self.nav_frame = tk.Frame(self.pagination_frame)
        self.nav_frame.grid(row=0, column=6, sticky="e")

        self.prev_btn = tk.Button(
            self.nav_frame, text="< Previous", command=lambda: self.change_page(-1)
        )
        self.prev_btn.pack(side="left")

        self.page_info = tk.Label(self.nav_frame, text="Page 1/1")
        self.page_info.pack(side="left", padx=5)

        self.next_btn = tk.Button(
            self.nav_frame, text="Next >", command=lambda: self.change_page(1)
        )
        self.next_btn.pack(side="left")

        # Canvas and Scrollbar for Orders
        self.orders_canvas = tk.Canvas(self.frame_orders_manager)
        self.orders_canvas.grid(row=2, column=0, sticky="nsew", padx=5, pady=5)
        self.orders_scrollbar = tk.Scrollbar(
            self.frame_orders_manager,
            orient=tk.VERTICAL,
            command=self.orders_canvas.yview,
        )
        self.orders_scrollbar.grid(row=2, column=1, sticky="ns", pady=5)
        self.orders_canvas.configure(yscrollcommand=self.orders_scrollbar.set)

        self.orders_inner_frame = tk.Frame(self.orders_canvas)
        self.orders_canvas.create_window(
            (0, 0), window=self.orders_inner_frame, anchor="nw"
        )

        self.orders_inner_frame.bind(
            "<Configure>",
            lambda e: self.orders_canvas.configure(
                scrollregion=self.orders_canvas.bbox("all")
            ),
        )
        self.orders_canvas.yview_moveto(0)
        self.orders_canvas.bind(
            "<Enter>",
            lambda e: self.orders_canvas.bind_all("<MouseWheel>", self._on_mousewheel),
        )
        self.orders_canvas.bind(
            "<Leave>", lambda e: self.orders_canvas.unbind_all("<MouseWheel>")
        )

        self.order_vars = {}
        self.order_checkbuttons = {}

        # Buttons below the canvas
        self.btn_frame = tk.Frame(self.frame_orders_manager)
        self.btn_frame.grid(row=3, column=0, columnspan=2, sticky="ew", padx=5, pady=5)
        self.btn_frame.columnconfigure((0, 1, 2), weight=1)

        self.btn_send_workplace = tk.Button(
            self.btn_frame, text="Send to Workplace", command=self.send_to_workplace
        )
        self.btn_send_workplace.grid(row=0, column=0, sticky="ew", padx=5)
        ToolTip(self.btn_send_workplace, "Send selected orders to workplace")

        self.btn_remove_workplace_orders = tk.Button(
            self.btn_frame,
            text="Remove Tests from Selected Orders",
            command=self.remove_workplace_orders_selected,
        )
        self.btn_remove_workplace_orders.grid(row=0, column=1, sticky="ew", padx=5)
        ToolTip(
            self.btn_remove_workplace_orders, "Remove selected orders from workplace"
        )

        self.btn_clear_workplace = tk.Button(
            self.btn_frame, text="Clear Workplace", command=self.clear_workplace
        )
        self.btn_clear_workplace.grid(row=0, column=2, sticky="ew", padx=5)
        ToolTip(self.btn_clear_workplace, "Clear all tests from workplace")

        # Workplace (right)
        self.workplace_frame = tk.Frame(
            self.root, relief="groove", bd=2, padx=10, pady=10
        )
        self.workplace_frame.grid(row=0, column=1, sticky="nsew")
        self.workplace_frame.columnconfigure(0, weight=1)
        self.workplace_frame.rowconfigure(2, weight=1)

        # Frame for Report and Close buttons aligned to the right at the top of the Workplace
        btn_report_frame = tk.Frame(self.workplace_frame)
        btn_report_frame.grid(row=0, column=0, sticky="ew")
        btn_report_frame.columnconfigure(0, weight=1)

        self.btn_report = tk.Button(
            btn_report_frame,
            text="Report",
            command=self.show_report,
            font=("Helvetica", 12, "bold"),
            bg="#4682b4",
            fg="white",
            padx=15,
            pady=10,
        )
        self.btn_report.grid(row=0, column=1, sticky="e", padx=(0, 5))
        ToolTip(self.btn_report, "Generate a report of workplace tests")

        self.btn_close_main = tk.Button(
            btn_report_frame,
            text="Close",
            command=self.close_application,
            font=("Helvetica", 12, "bold"),
            bg="#d9534f",
            fg="white",
            padx=15,
            pady=10,
        )
        self.btn_close_main.grid(row=0, column=2, sticky="e")
        self.btn_close_main.bind(
            "<Enter>", lambda e: self.btn_close_main.config(bg="#e57373")
        )
        self.btn_close_main.bind(
            "<Leave>", lambda e: self.btn_close_main.config(bg="#d9534f")
        )
        ToolTip(self.btn_close_main, "Close the application")

        self.workplace_title = tk.Label(
            self.workplace_frame, text="Workplace", font=("Helvetica", 14, "bold")
        )
        self.workplace_title.grid(row=0, column=0, sticky="w", pady=(0, 10))

        # Temperature and Limiter Filters
        filter_temp_frame = tk.Frame(self.workplace_frame)
        filter_temp_frame.grid(row=1, column=0, sticky="w", pady=5)

        tk.Label(filter_temp_frame, text="Filter Temperature:").pack(
            side="left", padx=(0, 5)
        )
        self.filter_temperature_var = tk.StringVar()
        self.filter_temperature_combobox = ttk.Combobox(
            filter_temp_frame,
            textvariable=self.filter_temperature_var,
            state="readonly",
            values=["All", "RT", "LT", "HT"],
            width=10,
        )
        self.filter_temperature_combobox.pack(side="left")
        self.filter_temperature_combobox.set("All")

        tk.Label(filter_temp_frame, text="Limiter:").pack(side="left", padx=(10, 5))
        self.limit_var = tk.StringVar()
        limit_values = ["All"] + [str(i) for i in range(0, 201, 1)]
        self.limit_combobox = ttk.Combobox(
            filter_temp_frame,
            textvariable=self.limit_var,
            state="readonly",
            values=limit_values,
            width=8,
        )
        self.limit_combobox.pack(side="left")
        self.limit_combobox.set("All")

        self.btn_apply_filters = tk.Button(
            filter_temp_frame, text="Apply Filters", command=self.apply_filters
        )
        self.btn_apply_filters.pack(side="left", padx=10)
        ToolTip(self.btn_apply_filters, "Apply temperature and limiter filters")

        # Results List with Scrollbars
        self.frame_results = tk.Frame(self.workplace_frame)
        self.frame_results.grid(row=2, column=0, sticky="nsew", pady=5)
        self.frame_results.columnconfigure(0, weight=1)
        self.frame_results.rowconfigure(0, weight=1)

        separator = tk.Frame(self.workplace_frame, height=2, bd=1, relief="sunken")
        separator.grid(row=3, column=0, sticky="ew", pady=(5, 5))

        self.counter_frame = tk.Frame(self.workplace_frame, pady=8)
        self.counter_frame.grid(row=4, column=0, sticky="ew")
        self.counter_frame.columnconfigure((0, 1, 2, 3), weight=1)

        self.label_rt = tk.Label(
            self.counter_frame,
            text="RT: 0",
            font=("Helvetica", 11, "bold"),
            fg="#008800",
        )
        self.label_rt.grid(row=0, column=0, sticky="ew", padx=8)

        self.label_lt = tk.Label(
            self.counter_frame,
            text="LT: 0",
            font=("Helvetica", 11, "bold"),
            fg="#0055cc",
        )
        self.label_lt.grid(row=0, column=1, sticky="ew", padx=8)

        self.label_ht = tk.Label(
            self.counter_frame,
            text="HT: 0",
            font=("Helvetica", 11, "bold"),
            fg="#cc5500",
        )
        self.label_ht.grid(row=0, column=2, sticky="ew", padx=8)

        self.label_total = tk.Label(
            self.counter_frame,
            text="Total: 0",
            font=("Helvetica", 11, "bold"),
            fg="#222222",
        )
        self.label_total.grid(row=0, column=3, sticky="ew", padx=8)

        self.scrollbar_y = tk.Scrollbar(self.frame_results, orient=tk.VERTICAL)
        self.scrollbar_x = tk.Scrollbar(self.frame_results, orient=tk.HORIZONTAL)

        self.list_results = tk.Listbox(
            self.frame_results,
            width=120,
            height=25,
            yscrollcommand=self.scrollbar_y.set,
            xscrollcommand=self.scrollbar_x.set,
        )
        self.list_results.grid(row=0, column=0, sticky="nsew")

        self.scrollbar_y.config(command=self.list_results.yview)
        self.scrollbar_y.grid(row=0, column=1, sticky="ns")

        self.scrollbar_x.config(command=self.list_results.xview)
        self.scrollbar_x.grid(row=1, column=0, sticky="ew")

        # Initializations
        self.json_file = "Data.json"
        # Chama o backup automático ao inicializar
        self.create_daily_backup()
        self.excel_folder = r"H:\TEAMS\Inflator_Lab\0_Evaluations\vi"
        self.workplace_data = []
        self.filtered_workplace_data = None

        self.update_orders_list()
        self.root.mainloop()

    def close_application(self):
        try:
            self.root.destroy()
            import sys

            sys.exit(0)
        except Exception as e:
            print(f"Error closing application: {str(e)}\n{traceback.format_exc()}")
            raise

    def _on_mousewheel(self, event):
        self.orders_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def create_daily_backup(self):
        try:
            # Verifica se o arquivo Data.json existe
            if not os.path.exists(self.json_file):
                print(f"Backup skipped: {self.json_file} not found.")
                return

            # Obtém a pasta do arquivo Data.json
            base_dir = os.path.dirname(self.json_file) or "."
            backup_dir = os.path.join(base_dir, "Backup")

            # Cria a pasta Backup se não existir
            if not os.path.exists(backup_dir):
                os.makedirs(backup_dir)
                print(f"Created Backup directory: {backup_dir}")

            # Obtém a data atual (YYYYMMDD)
            current_date = datetime.now().strftime("%Y%m%d")

            # Verifica se já existe um backup para a data atual
            backup_pattern = os.path.join(backup_dir, f"Data_{current_date}_*.json")
            existing_backups = glob.glob(backup_pattern)

            if existing_backups:
                print(
                    f"Backup already exists for {current_date}: {existing_backups[0]}"
                )
                return

            # Cria o nome do backup com data e hora
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"Data_{timestamp}.json"
            backup_path = os.path.join(backup_dir, backup_filename)

            # Copia o arquivo Data.json para a pasta Backup
            shutil.copy2(self.json_file, backup_path)
            print(f"Backup created: {backup_path}")

        except Exception as e:
            error_message = f"Error creating backup: {str(e)}"
            print(error_message)

    def process_orders(self):
        try:
            orders_input = self.entry_orders.get().strip()
            if not orders_input:
                messagebox.showerror("Error", "Please enter the order numbers.")
                return

            orders = [order.strip() for order in orders_input.split(",")]

            if os.path.exists(self.json_file):
                with open(self.json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
            else:
                data = {}

            files_to_process = []
            for order in orders:
                for file_name in os.listdir(self.excel_folder):
                    if file_name.startswith(order) and file_name.endswith(".xlsx"):
                        files_to_process.append(
                            os.path.join(self.excel_folder, file_name)
                        )

            if not files_to_process:
                messagebox.showerror(
                    "Error", "No Excel files found for the provided orders."
                )
                return

            for file in files_to_process:
                self.process_excel(file, data)

            with open(self.json_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            self.entry_orders.delete(0, tk.END)
            self.status_label.config(text="JSON database updated successfully!")
            messagebox.showinfo("Success", "Excel files processed and JSON updated!")
            self.update_orders_list()
        except Exception as e:
            messagebox.showerror("Error", f"Error processing orders: {str(e)}")

    def process_excel(self, file_path, data):
        wb = load_workbook(file_path, data_only=True)
        current_version = None
        current_order = None
        for sheet_name in wb.sheetnames:
            if "minus" in sheet_name.lower():
                temp_type = "LT"
            elif "rt" in sheet_name.lower():
                temp_type = "RT"
            elif "plus" in sheet_name.lower():
                temp_type = "HT"
            else:
                continue

            if "datenblatt" in sheet_name.lower():
                current_version, current_order = self.process_datenblatt(
                    wb[sheet_name], temp_type, data
                )
            elif "grafik" in sheet_name.lower() and current_version and current_order:
                self.process_grafik(
                    wb[sheet_name], temp_type, data, current_version, current_order
                )

    def process_datenblatt(self, sheet, temp_type, data):
        inflator_type = self.clean_value(sheet["U1"].value)
        version = "V" + inflator_type.split("V")[-1]

        test_order = self.clean_value(sheet["J4"].value)
        production_order = self.clean_value(sheet["J3"].value)
        propellant_lot_number = self.clean_value(sheet["S3"].value)
        test_date = self.parse_date(sheet["C4"].value)
        temperature_c = self.clean_value(sheet["C10"].value)

        if version not in data:
            data[version] = {}

        if test_order not in data[version]:
            data[version][test_order] = {
                "metadata": {
                    "production_order": production_order,
                    "propellant_lot_number": propellant_lot_number,
                    "test_date": test_date,
                },
                "temperatures": {},
            }

        if temp_type not in data[version][test_order]["temperatures"]:
            data[version][test_order]["temperatures"][temp_type] = {
                "temperature_c": float(temperature_c) if temperature_c else None,
                "tests": [],
            }

        tests = []
        seen_tests = set()
        for row in sheet.iter_rows(min_row=10, values_only=True):
            if row[0] and str(row[0]).strip().isdigit():
                test_no = self.clean_value(row[0])
                inflator_no = self.clean_value(row[1])
                if test_no and inflator_no and test_no not in seen_tests:
                    tests.append(
                        {"test_no": int(test_no), "inflator_no": int(inflator_no)}
                    )
                    seen_tests.add(test_no)

        data[version][test_order]["temperatures"][temp_type]["tests"] = tests
        return version, test_order

    def process_grafik(self, sheet, temp_type, data, current_version, current_order):
        valid_columns = []
        limits = {"maximums": {}, "minimums": {}}
        for col in range(self.MIN_COLUMN, self.MAX_COLUMN):
            min_val = self.clean_value(
                sheet.cell(row=self.MIN_LIMIT_ROW, column=col).value
            )
            max_val = self.clean_value(
                sheet.cell(row=self.MAX_LIMIT_ROW, column=col).value
            )
            if min_val or max_val:
                valid_columns.append(col)
                ms = col - 2
                if min_val is not None:
                    try:
                        limits["minimums"][str(ms)] = float(min_val)
                    except ValueError:
                        pass
                if max_val is not None:
                    try:
                        limits["maximums"][str(ms)] = float(max_val)
                    except ValueError:
                        pass

        inflator_nos = [
            test["inflator_no"]
            for test in data[current_version][current_order]["temperatures"][temp_type][
                "tests"
            ]
        ]

        pressure_data = []
        blank_line_count = 0
        row_idx = self.PRESSURE_DATA_START_ROW

        for inflator_no in inflator_nos:
            is_blank = True
            pressures = {}
            for col in valid_columns:
                pressure = self.clean_value(sheet.cell(row=row_idx, column=col).value)
                if pressure is not None:
                    try:
                        pressures[str(col - 2)] = float(pressure)
                        is_blank = False
                    except ValueError:
                        continue

            if is_blank:
                blank_line_count += 1
                if blank_line_count >= 2:
                    break
            else:
                blank_line_count = 0
                if pressures:
                    pressure_data.append(
                        {"inflator_no": inflator_no, "pressures": pressures}
                    )

            row_idx += 1

        data[current_version][current_order]["temperatures"][temp_type][
            "pressure_data"
        ] = pressure_data
        data[current_version][current_order]["temperatures"][temp_type][
            "limits"
        ] = limits

    def update_orders_list(self):
        for widget in self.orders_inner_frame.winfo_children():
            widget.destroy()
        self.order_checkbuttons.clear()

        if not os.path.exists(self.json_file):
            return

        try:
            with open(self.json_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            versions = sorted(data.keys())
            self.version_combobox["values"] = ["All"] + versions
            current = self.version_combobox.get()
            if current not in self.version_combobox["values"]:
                self.version_combobox.set("All")
            selected_version = self.version_combobox.get()

            orders_list = []
            for version, orders in data.items():
                if (
                    selected_version
                    and selected_version.lower() != "all"
                    and version != selected_version
                ):
                    continue
                for order, details in orders.items():
                    test_date = details["metadata"].get("test_date", "0000-00-00")
                    orders_list.append((version, order, test_date))

            def parse_date_safe(date_str):
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except Exception:
                    return datetime.min

            orders_list.sort(key=lambda x: parse_date_safe(x[2]), reverse=True)

            total_orders = len(orders_list)
            self.total_pages = max(
                1, (total_orders + self.orders_per_page - 1) // self.orders_per_page
            )
            if not hasattr(self, "current_page") or self.current_page < 1:
                self.current_page = 1
            if self.current_page > self.total_pages:
                self.current_page = self.total_pages

            start_idx = (self.current_page - 1) * self.orders_per_page
            end_idx = start_idx + self.orders_per_page
            paginated_orders = orders_list[start_idx:end_idx]

            for idx, (version, order, test_date) in enumerate(
                paginated_orders, start=start_idx + 1
            ):
                # Reutiliza BooleanVar existente ou cria novo
                key = (version, order)
                if key not in self.order_vars:
                    self.order_vars[key] = tk.BooleanVar()
                var = self.order_vars[key]
                display_text = (
                    f"{idx}. Version: {version}, Order: {order}, Date: {test_date}"
                )

                row_frame = tk.Frame(self.orders_inner_frame)
                row_frame.grid(
                    row=idx - start_idx, column=0, sticky="w", padx=5, pady=2
                )

                chk = tk.Checkbutton(
                    row_frame,
                    text=display_text,
                    variable=var,
                    anchor="w",
                    width=45,
                )
                chk.pack(side=tk.LEFT)

                btn_view = tk.Button(
                    row_frame,
                    text="     👁️",
                    width=3,
                    command=lambda v=version, o=order: self.show_metadata_popup(v, o),
                )
                btn_view.pack(side=tk.LEFT, padx=(10, 0))

                self.order_checkbuttons[key] = chk

            self.orders_canvas.configure(scrollregion=self.orders_canvas.bbox("all"))
            self.orders_canvas.yview_moveto(0)

            self.page_info.config(text=f"Page {self.current_page}/{self.total_pages}")

            self.prev_btn.config(
                state=tk.NORMAL if self.current_page > 1 else tk.DISABLED
            )
            self.next_btn.config(
                state=tk.NORMAL if self.current_page < self.total_pages else tk.DISABLED
            )

            # Atualiza o estado do Select All baseado na página atual
            current_page_orders = [
                (version, order) for version, order, _ in paginated_orders
            ]
            all_selected = all(
                self.order_vars.get(key, tk.BooleanVar(value=False)).get()
                for key in current_page_orders
            )
            self.select_all_var.set(all_selected)

        except Exception as e:
            messagebox.showerror("Error", f"Error loading orders: {str(e)}")

    def show_metadata_popup(self, version, order):
        if not os.path.exists(self.json_file):
            messagebox.showerror("Error", "Database not found.")
            return

        with open(self.json_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        if version not in data or order not in data[version]:
            messagebox.showerror("Error", "Order not found in the database.")
            return

        metadata = data[version][order].get("metadata", {})
        temperatures = data[version][order].get("temperatures", {})

        info = ""
        for k, v in metadata.items():
            info += f"{k}: {v}\n"
        info += "\nTemperatures (°C):\n"
        for tipo, tdata in temperatures.items():
            info += f"  {tipo}: {tdata.get('temperature_c', 'N/A')}\n"

        popup = tk.Toplevel(self.root)
        popup.title(f"Metadata - {order}")
        popup.geometry("350x250")
        popup.resizable(False, False)
        tk.Label(
            popup,
            text=f"Metadata for Order {order} ({version})",
            font=("Arial", 11, "bold"),
        ).pack(pady=8)
        text = tk.Text(popup, width=40, height=10, wrap="word")
        text.insert("1.0", info)
        text.config(state="disabled")
        text.pack(padx=8, pady=8)
        tk.Button(popup, text="Close", command=popup.destroy).pack(pady=5)

    def update_items_per_page(self, event=None):
        try:
            self.orders_per_page = int(self.page_selector.get())
            self.current_page = 1
            self.update_orders_list()
        except Exception as e:
            messagebox.showerror("Error", f"Error updating items per page: {str(e)}")

    def on_items_per_page_changed(self, event=None):
        try:
            self.orders_per_page = int(self.page_selector.get())
        except Exception:
            self.orders_per_page = 10
        self.current_page = 1
        self.update_orders_list()

    def change_page(self, delta):
        new_page = self.current_page + delta
        if 1 <= new_page <= self.total_pages:
            self.current_page = new_page
            self.update_orders_list()

    def toggle_select_all(self):
        state = self.select_all_var.get()
        start_idx = (self.current_page - 1) * self.orders_per_page
        end_idx = start_idx + self.orders_per_page

        try:
            with open(self.json_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            orders_list = []
            selected_version = self.version_combobox.get()
            for version, orders in data.items():
                if (
                    selected_version
                    and selected_version.lower() != "all"
                    and version != selected_version
                ):
                    continue
                for order, details in orders.items():
                    test_date = details["metadata"].get("test_date", "0000-00-00")
                    orders_list.append((version, order, test_date))

            def parse_date_safe(date_str):
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except Exception:
                    return datetime.min

            orders_list.sort(key=lambda x: parse_date_safe(x[2]), reverse=True)
            paginated_orders = orders_list[start_idx:end_idx]

            for version, order, _ in paginated_orders:
                key = (version, order)
                if key in self.order_vars:
                    self.order_vars[key].set(state)

        except Exception as e:
            messagebox.showerror("Error", f"Error toggling select all: {str(e)}")

    def on_version_filter(self, event=None):
        self.current_page = 1
        # Limpa todas as seleções em order_vars
        for var in self.order_vars.values():
            var.set(False)
        self.select_all_var.set(False)  # Desmarca o "Select All"
        self.update_orders_list()

    def remove_orders_by_input(self):
        try:
            orders_input = self.entry_orders.get().strip()
            if not orders_input:
                messagebox.showerror("Error", "Please enter the order numbers.")
                return

            orders_to_remove = [order.strip() for order in orders_input.split(",")]

            if not os.path.exists(self.json_file):
                messagebox.showerror("Error", "No database found.")
                return

            with open(self.json_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            removed = []
            for version in list(data.keys()):
                for order in list(data[version].keys()):
                    if order in orders_to_remove:
                        del data[version][order]
                        removed.append(order)
                        if not data[version]:
                            del data[version]

            with open(self.json_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            if removed:
                msg = f"Orders removed successfully:\n{', '.join(removed)}"
                self.status_label.config(text=msg)
                messagebox.showinfo("Success", msg)
            else:
                messagebox.showwarning("Warning", "No matching orders found.")

            self.entry_orders.delete(0, tk.END)
            self.update_orders_list()

        except Exception as e:
            messagebox.showerror("Error", f"Error removing orders: {str(e)}")

    def send_to_workplace(self):
        selected_orders = [
            (version, order)
            for (version, order), var in self.order_vars.items()
            if var.get()
        ]
        if not selected_orders:
            messagebox.showwarning(
                "Warning", "No orders selected to send to workplace."
            )
            return

        try:
            with open(self.json_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            versions = set(version for version, order in selected_orders)
            if len(versions) > 1:
                messagebox.showerror("Error", "Select tests from only one version.")
                return

            if (
                self.workplace_data
                and self.workplace_data[0]["version"] not in versions
            ):
                messagebox.showerror(
                    "Error",
                    "Workplace already contains tests from another version. Clear workplace first.",
                )
                return

            existing_keys = set(
                (
                    reg["test_no"],
                    reg["inflator_no"],
                    reg["type"],
                    reg["version"],
                    reg["order"],
                )
                for reg in self.workplace_data
            )

            new_workplace_data = []
            duplicates_skipped = 0

            for version, order in selected_orders:
                if version in data and order in data[version]:
                    details = data[version][order]
                    metadata = details.get("metadata", {})
                    test_date = metadata.get("test_date", "0000-00-00")
                    temperatures = details.get("temperatures", {})
                    for temp_type in ["RT", "LT", "HT"]:
                        if temp_type not in temperatures:
                            continue
                        temp_data = temperatures[temp_type]
                        temperature_c = temp_data.get("temperature_c", "N/A")
                        tests = temp_data.get("tests", [])
                        pressure_data = temp_data.get("pressure_data", [])
                        pressure_map = {
                            item["inflator_no"]: item["pressures"]
                            for item in pressure_data
                        }
                        for test in tests:
                            test_no = test.get("test_no", "N/A")
                            inflator_no = test.get("inflator_no", "N/A")
                            key = (test_no, inflator_no, temp_type, version, order)
                            if key in existing_keys:
                                duplicates_skipped += 1
                                continue
                            new_workplace_data.append(
                                {
                                    "test_no": test_no,
                                    "inflator_no": inflator_no,
                                    "temperature_c": temperature_c,
                                    "type": temp_type,
                                    "version": version,
                                    "order": order,
                                    "test_date": test_date,
                                    "pressures": pressure_map.get(inflator_no, {}),
                                }
                            )
                            existing_keys.add(key)

            self.workplace_data.extend(new_workplace_data)

            def parse_date_safe(date_str):
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except:
                    return datetime(1900, 1, 1)

            self.workplace_data.sort(
                key=lambda x: parse_date_safe(x.get("test_date", "1900-01-01")),
                reverse=True,
            )

            self.update_workplace_display()
            self.update_workplace_counters()

            msg = f"Tests sent to workplace successfully! Added {len(new_workplace_data)} records."
            if duplicates_skipped > 0:
                msg += f"\n{duplicates_skipped} duplicate test(s) were ignored."
            messagebox.showinfo("Success", msg)

        except Exception as e:
            messagebox.showerror("Error", f"Error sending tests to workplace: {str(e)}")

    def update_workplace_display(self):
        self.list_results.delete(0, tk.END)
        header = "Test | Inflator | Temperature | Type | Version | Order | Date"
        self.list_results.insert(tk.END, header)
        self.list_results.insert(tk.END, "-" * len(header))

        for reg in self.workplace_data:
            line = f"{reg['test_no']} | {reg['inflator_no']} | {reg['temperature_c']}°C | {reg['type']} | {reg['version']} | {reg['order']} | {reg['test_date']}"
            if reg["pressures"]:
                line += " | Pressure data available"
            else:
                line += " | No pressure data"
            self.list_results.insert(tk.END, line)
        self.update_workplace_counters()

    def apply_filters(self):
        temp_filter = self.filter_temperature_var.get()
        limit_filter = self.limit_var.get()

        self.list_results.delete(0, tk.END)
        header = "Test | Inflator | Temperature | Type | Version | Order | Date"
        self.list_results.insert(tk.END, header)
        self.list_results.insert(tk.END, "-" * len(header))

        versions = {reg["version"] for reg in self.workplace_data}
        if len(versions) > 1:
            messagebox.showerror(
                "Error",
                "Workplace contains mixed versions! Clear before applying filters.",
            )
            return

        if temp_filter == "All" and limit_filter == "All":
            self.filtered_workplace_data = None
            for reg in self.workplace_data:
                line = f"{reg['test_no']} | {reg['inflator_no']} | {reg['temperature_c']}°C | {reg['type']} | {reg['version']} | {reg['order']} | {reg['test_date']}"
                if reg["pressures"]:
                    line += " | Pressure data available"
                else:
                    line += " | No pressure data"
                self.list_results.insert(tk.END, line)
            self.update_workplace_counters()
            return

        filtered_data = []
        for reg in self.workplace_data:
            if temp_filter != "All" and reg["type"] != temp_filter:
                continue
            filtered_data.append(reg)

        if limit_filter != "All":
            limit = int(limit_filter)
            filtered_data = filtered_data[:limit]

        for reg in filtered_data:
            line = f"{reg['test_no']} | {reg['inflator_no']} | {reg['temperature_c']}°C | {reg['type']} | {reg['version']} | {reg['order']} | {reg['test_date']}"
            if reg["pressures"]:
                line += " | Pressure data available"
            else:
                line += " | No pressure data"
            self.list_results.insert(tk.END, line)
        self.filtered_workplace_data = filtered_data
        self.update_workplace_counters(filtered_data)

    def remove_workplace_orders_selected(self):
        selected_orders = {
            (version, order)
            for (version, order), var in self.order_vars.items()
            if var.get()
        }
        if not selected_orders:
            messagebox.showwarning(
                "Warning", "No orders selected to remove from Workplace."
            )
            return

        before = len(self.workplace_data)
        self.workplace_data = [
            reg
            for reg in self.workplace_data
            if (reg["version"], reg["order"]) not in selected_orders
        ]
        after = len(self.workplace_data)
        self.update_workplace_display()
        self.update_workplace_counters()
        messagebox.showinfo(
            "Success", f"{before - after} records removed from Workplace."
        )

    def update_workplace_counters(self, data=None):
        if data is None:
            data = self.workplace_data
        rt = lt = ht = 0
        for item in data:
            temp = item.get("type")
            if temp == "RT":
                rt += 1
            elif temp == "LT":
                lt += 1
            elif temp == "HT":
                ht += 1
        total = rt + lt + ht
        self.label_rt.config(text=f"RT: {rt}")
        self.label_lt.config(text=f"LT: {lt}")
        self.label_ht.config(text=f"HT: {ht}")
        self.label_total.config(text=f"Total: {total}")

    def clear_workplace(self):
        self.workplace_data.clear()
        self.list_results.delete(0, tk.END)
        messagebox.showinfo("Success", "Workplace cleared successfully.")
        self.update_workplace_counters()

    def clean_value(self, value):
        if value is None:
            return None
        if isinstance(value, str):
            return value.strip()
        return value

    def parse_date(self, date_value):
        if isinstance(date_value, datetime):
            return date_value.strftime("%Y-%m-%d")
        if isinstance(date_value, str):
            try:
                dt = datetime.strptime(date_value, "%d.%m.%Y")
                return dt.strftime("%Y-%m-%d")
            except Exception:
                return date_value
        return None

    def export_database_to_excel(self):
        try:
            if not os.path.exists(self.json_file):
                messagebox.showerror("Error", "Database file not found.")
                return

            with open(self.json_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            if not data:
                messagebox.showwarning("Warning", "Database is empty.")
                return

            wb = Workbook()
            wb.remove(wb.active)

            # Estilos
            center_alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            double_border = Border(
                left=Side(style="double"),
                right=Side(style="double"),
                top=Side(style="double"),
                bottom=Side(style="double"),
            )
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            title_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
            metadata_font = Font(name="Calibri", size=10, bold=True)
            header_font = Font(name="Calibri", size=10, bold=True)
            data_font = Font(name="Calibri", size=10)
            note_font = Font(name="Calibri", size=9, italic=True)
            title_fill = PatternFill(
                start_color="4682B4", end_color="4682B4", fill_type="solid"
            )
            metadata_fill = PatternFill(
                start_color="E6F0FA", end_color="E6F0FA", fill_type="solid"
            )
            rt_fill = PatternFill(
                start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
            )
            lt_fill = PatternFill(
                start_color="CCE6FF", end_color="CCE6FF", fill_type="solid"
            )
            ht_fill = PatternFill(
                start_color="FFE6CC", end_color="FFE6CC", fill_type="solid"
            )
            alt_fill = PatternFill(
                start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
            )
            warning_fill = PatternFill(
                start_color="FF9999", end_color="FF9999", fill_type="solid"
            )
            low_fill = PatternFill(
                start_color="99CCFF", end_color="99CCFF", fill_type="solid"
            )

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Data_{timestamp}.xlsx"

            # Aba de Resumo
            ws_summary = wb.create_sheet(title="Summary", index=0)
            ws_summary.append(["Ballistic Tests Database Summary"])
            ws_summary.append(
                [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
            )
            ws_summary.append([])

            summary_stats = {
                "total_orders": 0,
                "total_tests": 0,
                "tests_by_temp": {"RT": 0, "LT": 0, "HT": 0},
            }
            for version in data:
                summary_stats["total_orders"] += len(data[version])
                for order, order_data in data[version].items():
                    for temp_type, temp_data in order_data["temperatures"].items():
                        summary_stats["tests_by_temp"][temp_type] += len(
                            temp_data.get("tests", [])
                        )
                        summary_stats["total_tests"] += len(temp_data.get("tests", []))

            ws_summary.append(["Total Orders", summary_stats["total_orders"]])
            ws_summary.append(["Total Tests", summary_stats["total_tests"]])
            ws_summary.append(["RT Tests", summary_stats["tests_by_temp"]["RT"]])
            ws_summary.append(["LT Tests", summary_stats["tests_by_temp"]["LT"]])
            ws_summary.append(["HT Tests", summary_stats["tests_by_temp"]["HT"]])
            ws_summary.append([])
            ws_summary.append(
                ["Note: Detailed data is available in version-specific sheets."]
            )

            for row in ws_summary["A1:A2"]:
                for cell in row:
                    cell.font = title_font
                    cell.fill = title_fill
                    cell.alignment = center_alignment
                    cell.border = double_border
            for row in ws_summary["A4:B8"]:
                for cell in row:
                    cell.font = metadata_font
                    cell.fill = metadata_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border
            ws_summary["A10"].font = note_font
            ws_summary["A10"].alignment = center_alignment

            # Ajusta largura e altura da aba de resumo
            column_widths = {}
            for row in ws_summary.iter_rows():
                for cell in row:
                    if cell.value:
                        col_letter = cell.column_letter
                        current_width = column_widths.get(col_letter, 0)
                        cell_len = len(str(cell.value)) + 2  # Margem extra
                        column_widths[col_letter] = max(current_width, cell_len)
            for col_letter, width in column_widths.items():
                adjusted_width = min(max(width * 1.1, 10), 50)
                ws_summary.column_dimensions[col_letter].width = adjusted_width

            for row in ws_summary.iter_rows():
                max_height = 15  # Altura mínima
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        lines = cell.value.count("\n") + 1
                        width = column_widths.get(cell.column_letter, 10)
                        chars_per_line = max(
                            len(line) for line in cell.value.split("\n")
                        )
                        if chars_per_line > width:
                            lines += chars_per_line // width
                        max_height = max(max_height, lines * 15)
                ws_summary.row_dimensions[row[0].row].height = max_height

            for version in sorted(data.keys()):
                ws = wb.create_sheet(title=version)
                total_orders = len(data[version])
                total_tests = 0

                # Metadados
                ws.append(["Ballistic Tests Database 🌐"])
                ws.append([f"Version: {version}"])
                ws.append(
                    [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                )
                ws.append([f"Total Orders: {total_orders}"])
                ws.append([])

                for row in ws["A1:A4"]:
                    for cell in row:
                        cell.font = title_font
                        cell.fill = title_fill
                        cell.alignment = center_alignment
                        cell.border = double_border

                # Cabeçalhos principais
                main_headers = [
                    "Order Number 📋",
                    "Production Order 🏭",
                    "Propellant Lot Number 🧪",
                    "Test Date 📅",
                    "Test Number 🔢",
                    "Inflator Number ⚙️",
                ]
                ws.append(main_headers)
                header_row = 6
                for col_idx, header in enumerate(main_headers, 1):
                    cell = ws.cell(row=header_row, column=col_idx)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = metadata_fill
                    cell.border = thin_border
                    cell.alignment = center_alignment

                # Coleta de pontos de tempo (ms) e limites por temperatura
                ms_points_by_temp = {"RT": set(), "LT": set(), "HT": set()}
                limits_by_temp = {
                    "RT": {"max": {}, "min": {}},
                    "LT": {"max": {}, "min": {}},
                    "HT": {"max": {}, "min": {}},
                }
                for order, order_data in data[version].items():
                    for temp_type, temp_data in order_data["temperatures"].items():
                        for pressure_entry in temp_data.get("pressure_data", []):
                            ms_points_by_temp[temp_type].update(
                                str(key) for key in pressure_entry["pressures"].keys()
                            )
                        limits = temp_data.get("limits", {})
                        limits_by_temp[temp_type]["max"].update(
                            limits.get("maximums", {})
                        )
                        limits_by_temp[temp_type]["min"].update(
                            limits.get("minimums", {})
                        )

                # Organiza os pontos de tempo
                for temp in ms_points_by_temp:
                    ms_points_by_temp[temp] = sorted(ms_points_by_temp[temp], key=int)

                # Cabeçalhos de temperatura e tempos
                temp_headers_row = header_row + 1
                col_idx = len(main_headers) + 1
                for temp in ["RT", "LT", "HT"]:
                    if not ms_points_by_temp[temp]:
                        continue
                    ms_points = ms_points_by_temp[temp]
                    ws.cell(row=temp_headers_row, column=col_idx).value = (
                        f"{temp} Data 🌡️"
                    )
                    ws.merge_cells(
                        start_row=temp_headers_row,
                        start_column=col_idx,
                        end_row=temp_headers_row,
                        end_column=col_idx + len(ms_points) - 1,
                    )
                    cell = ws.cell(row=temp_headers_row, column=col_idx)
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    cell.fill = (
                        rt_fill
                        if temp == "RT"
                        else lt_fill if temp == "LT" else ht_fill
                    )

                    for ms in ms_points:
                        cell = ws.cell(row=temp_headers_row + 1, column=col_idx)
                        cell.value = f"{ms} ms ⏱️"  # ms já é string
                        cell.font = header_font
                        cell.alignment = center_alignment
                        cell.border = thin_border
                        cell.fill = (
                            rt_fill
                            if temp == "RT"
                            else lt_fill if temp == "LT" else ht_fill
                        )
                        col_idx += 1

                # Ordena ordens por data (mais recente para mais antiga)
                def parse_date_safe(date_str):
                    try:
                        return datetime.strptime(date_str, "%Y-%m-%d")
                    except Exception:
                        return datetime.min

                orders_sorted = sorted(
                    data[version].items(),
                    key=lambda x: parse_date_safe(
                        x[1].get("metadata", {}).get("test_date", "0000-00-00")
                    ),
                    reverse=True,
                )

                # Dados
                row_idx = temp_headers_row + 2
                start_merge_row = row_idx
                for order, order_data in orders_sorted:
                    metadata = order_data.get("metadata", {})
                    production_order = metadata.get("production_order", "N/A")
                    propellant_lot = metadata.get("propellant_lot_number", "N/A")
                    test_date = metadata.get("test_date", "N/A")
                    temp_types = ["RT", "LT", "HT"]  # Ordem fixa: RT, LT, HT
                    tests_by_temp = {}
                    pressure_maps = {}

                    for temp_type in temp_types:
                        if temp_type not in order_data["temperatures"]:
                            continue
                        temp_data = order_data["temperatures"][temp_type]
                        temp_c = temp_data.get("temperature_c", "N/A")
                        tests = temp_data.get("tests", [])
                        pressure_data = temp_data.get("pressure_data", [])
                        pressure_maps[temp_type] = {
                            entry["inflator_no"]: entry["pressures"]
                            for entry in pressure_data
                        }
                        tests_by_temp[temp_type] = [
                            {
                                "test_no": test.get("test_no", "N/A"),
                                "inflator_no": test.get("inflator_no", "N/A"),
                                "temp_c": str(temp_c),
                            }
                            for test in tests
                        ]

                    # Conta total de testes para mesclagem
                    total_rows_for_order = sum(
                        len(tests_by_temp.get(temp, [])) for temp in temp_types
                    )
                    if total_rows_for_order == 0:
                        continue

                    for temp_type in temp_types:
                        if temp_type not in tests_by_temp:
                            continue
                        for test in tests_by_temp[temp_type]:
                            row = [
                                str(order),  # Garante que order seja string
                                str(production_order),
                                str(propellant_lot),
                                str(test_date),
                                str(test["test_no"]),
                                str(test["inflator_no"]),
                            ]
                            # Adiciona colunas de pressão para todas as temperaturas
                            for temp in ["RT", "LT", "HT"]:
                                ms_points = ms_points_by_temp[temp]
                                pressures = pressure_maps.get(temp, {}).get(
                                    test["inflator_no"], {}
                                )
                                if temp == temp_type:
                                    for ms in ms_points:
                                        val = pressures.get(str(ms), np.nan)
                                        formatted_val = (
                                            f"{val:.2f}" if not np.isnan(val) else "-"
                                        )
                                        row.append(formatted_val)
                                else:
                                    row.extend([""] * len(ms_points))
                            ws.append(row)
                            total_tests += 1

                            # Estiliza a linha e aplica formatação condicional
                            for col_idx in range(1, len(row) + 1):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.font = data_font
                                cell.border = thin_border
                                cell.alignment = center_alignment
                                if row_idx % 2 == 0:
                                    cell.fill = alt_fill
                                # Formatação condicional para valores de pressão
                                if col_idx > len(main_headers) and cell.value not in [
                                    "",
                                    "-",
                                ]:
                                    try:
                                        val = float(cell.value)
                                        ms_idx = col_idx - len(main_headers) - 1
                                        temp_idx = next(
                                            i
                                            for i, t in enumerate(["RT", "LT", "HT"])
                                            if ms_idx
                                            < sum(
                                                len(ms_points_by_temp[t])
                                                for t in ["RT", "LT", "HT"][: i + 1]
                                            )
                                        )
                                        temp = ["RT", "LT", "HT"][temp_idx]
                                        ms = ms_points_by_temp[temp][
                                            ms_idx
                                            - sum(
                                                len(ms_points_by_temp[t])
                                                for t in ["RT", "LT", "HT"][:temp_idx]
                                            )
                                        ]
                                        max_limit = limits_by_temp[temp]["max"].get(
                                            str(ms)
                                        )
                                        min_limit = limits_by_temp[temp]["min"].get(
                                            str(ms)
                                        )
                                        if max_limit is not None and val > float(
                                            max_limit
                                        ):
                                            cell.fill = warning_fill
                                        elif min_limit is not None and val < float(
                                            min_limit
                                        ):
                                            cell.fill = low_fill
                                    except (ValueError, IndexError):
                                        pass
                            row_idx += 1

                    # Mescla células para a ordem atual
                    if total_rows_for_order > 1:
                        for col in range(1, 5):
                            ws.merge_cells(
                                start_row=start_merge_row,
                                start_column=col,
                                end_row=start_merge_row + total_rows_for_order - 1,
                                end_column=col,
                            )
                    start_merge_row = row_idx

                # Total de testes
                ws.append([])
                ws.append([f"Total Tests: {total_tests}"])
                for cell in ws[f"A{row_idx + 1}:A{row_idx + 1}"]:
                    cell[0].font = metadata_font
                    cell[0].fill = metadata_fill
                    cell[0].alignment = center_alignment
                    cell[0].border = double_border

                # Notas de rodapé
                ws.append([])
                ws.append(["Notes:"])
                ws.append(
                    ["- Values in red indicate pressures above the maximum limit."]
                )
                ws.append(
                    ["- Values in blue indicate pressures below the minimum limit."]
                )
                ws.append(["- Use filters to sort or analyze data."])
                for row in ws[f"A{row_idx + 3}:A{row_idx + 6}"]:
                    for cell in row:
                        cell.font = note_font
                        cell.alignment = center_alignment
                        cell.border = thin_border

                # Ajusta largura das colunas
                column_widths = {}
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value:
                            col_letter = cell.column_letter
                            current_width = column_widths.get(col_letter, 0)
                            cell_len = len(str(cell.value)) + 2  # Margem extra
                            column_widths[col_letter] = max(current_width, cell_len)
                for col_letter, width in column_widths.items():
                    adjusted_width = min(max(width * 1.1, 10), 25)
                    ws.column_dimensions[col_letter].width = adjusted_width

                # Ajusta altura das linhas
                for row in ws.iter_rows():
                    row_number = row[0].row
                    max_height = 15  # Altura mínima padrão
                    height_scale = 15  # Escala padrão para linhas de dados

                    # Ajuste específico para linhas 1-4 (metadados) e linha 6 (cabeçalhos principais)
                    if 1 <= row_number <= 4:
                        max_height = 40  # Altura mínima maior para metadados
                        height_scale = 20  # Escala maior para melhor aparência
                    if row_number == 6:
                        max_height = 25  # Altura mínima para cabeçalhos
                        height_scale = 20  # Escala maior para melhor aparência

                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            lines = cell.value.count("\n") + 1
                            width = column_widths.get(cell.column_letter, 10)
                            chars_per_line = max(
                                len(line) for line in cell.value.split("\n")
                            )
                            if chars_per_line > width:
                                lines += chars_per_line // width
                            max_height = max(max_height, lines * height_scale)

                    ws.row_dimensions[row_number].height = max_height

                # Fixa cabeçalhos, incluindo Inflator Number
                ws.freeze_panes = "G9"

                # Adiciona filtros
                table_range = (
                    f"A{header_row}:{get_column_letter(ws.max_column)}{row_idx}"
                )
                tab = Table(displayName=f"Table_{version}", ref=table_range)
                style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                tab.tableStyleInfo = style
                ws.add_table(tab)

                # Protege a planilha
                ws.protection = SheetProtection(
                    sheet=True,
                    formatCells=False,
                    formatColumns=False,
                    formatRows=False,
                    insertRows=False,
                    insertColumns=False,
                    deleteRows=False,
                    deleteColumns=False,
                    sort=True,
                    autoFilter=True,
                )

            # Aba de instruções
            ws_instructions = wb.create_sheet(title="Instructions", index=1)
            ws_instructions.append(
                ["Instructions for Using the Ballistic Tests Database 📖"]
            )
            ws_instructions.append([])
            instructions = [
                "Welcome to the Ballistic Tests Database Export!",
                "This Excel file contains test data organized by version, with each version in a separate sheet.",
                "",
                "Key Features:",
                "- The 'Summary' sheet provides an overview of total orders and tests.",
                "- Each version sheet (e.g., V1, V2) contains detailed test data.",
                "- Data is grouped by Order Number, sorted by test date (most recent first).",
                "- Within each order, tests are ordered by temperature: RT (Room Temperature), LT (Low Temperature), HT (High Temperature).",
                "- Pressure values are listed in milliseconds (ms) under each temperature.",
                "- Headers are frozen for easy navigation (scroll while keeping headers visible).",
                "- Filters are enabled for sorting and analyzing data.",
                "",
                "Color Coding:",
                "- RT Headers: Light Green 🌡️",
                "- LT Headers: Light Blue 🌡️",
                "- HT Headers: Light Orange 🌡️",
                "- Red cells indicate pressures above the maximum limit 🚨",
                "- Blue cells indicate pressures below the minimum limit ❄️",
                "- Alternating rows have a light gray background for readability.",
                "",
                "Tips:",
                "- Use the 'Summary' sheet to get a quick overview.",
                "- Apply filters to explore specific orders or tests.",
                "- Protected sheets prevent accidental edits but allow filtering and sorting.",
                "- Contact the Inflator Lab team for support if needed.",
            ]
            for line in instructions:
                ws_instructions.append([line])
            for row in ws_instructions["A1:A1"]:
                for cell in row:
                    cell.font = title_font
                    cell.fill = title_fill
                    cell.alignment = center_alignment
                    cell.border = double_border
            for row in ws_instructions["A2:A{}".format(ws_instructions.max_row)]:
                for cell in row:
                    cell.font = data_font
                    cell.alignment = center_alignment
                    cell.border = thin_border

            # Ajusta largura e altura da aba de instruções
            column_widths = {}
            for row in ws_instructions.iter_rows():
                for cell in row:
                    if cell.value:
                        col_letter = cell.column_letter
                        current_width = column_widths.get(col_letter, 0)
                        cell_len = len(str(cell.value)) + 2  # Margem extra
                        column_widths[col_letter] = max(current_width, cell_len)
            for col_letter, width in column_widths.items():
                adjusted_width = min(max(width * 1.1, 10), 100)
                ws_instructions.column_dimensions[col_letter].width = adjusted_width

            for row in ws_instructions.iter_rows():
                max_height = 15  # Altura mínima
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        lines = cell.value.count("\n") + 1
                        width = column_widths.get(cell.column_letter, 10)
                        chars_per_line = max(
                            len(line) for line in cell.value.split("\n")
                        )
                        if chars_per_line > width:
                            lines += chars_per_line // width
                        max_height = max(max_height, lines * 15)
                ws_instructions.row_dimensions[row[0].row].height = max_height

            wb.save(filename)
            messagebox.showinfo("Success", f"Database exported to {filename}")

        except Exception as e:
            error_message = (
                f"Error exporting database: {str(e)}\n{traceback.format_exc()}"
            )
            print(error_message)
            messagebox.showerror("Error", f"Error exporting database: {str(e)}")

    def show_report(self):
        try:
            # Use filtered data if available, otherwise use all workplace data
            data_to_use = (
                self.filtered_workplace_data
                if self.filtered_workplace_data is not None
                else self.workplace_data
            )

            if not data_to_use:
                messagebox.showwarning(
                    "Warning",
                    "Workplace empty or no data after filtering. Add tests before generating the report.",
                )
                return

            # Group data by temperature
            data_by_temp = {}
            for reg in data_to_use:
                temp = reg["type"]
                if temp not in data_by_temp:
                    data_by_temp[temp] = []
                data_by_temp[temp].append(reg)

            # Create report window
            report_win = tk.Toplevel(self.root)
            report_win.title("Ballistic Tests Report")
            report_win.geometry("1000x700")
            report_win.minsize(800, 600)
            report_win.focus_set()
            report_win.grab_set()

            # Header
            header_frame = tk.Frame(report_win, bg="#fafafa")
            header_frame.pack(fill=tk.X, padx=5, pady=5)
            tk.Label(
                header_frame,
                text="Ballistic Tests Report",
                font=("Helvetica", 16, "bold"),
                bg="#fafafa",
            ).pack(side=tk.LEFT)
            tk.Label(
                header_frame,
                text=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                font=("Helvetica", 10),
                bg="#fafafa",
            ).pack(side=tk.RIGHT)

            # Frame for buttons in top-right corner
            btn_frame = tk.Frame(report_win, bg="#fafafa")
            btn_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

            # Frame with vertical scrollbar
            container = ttk.Frame(report_win)
            container.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
            container.columnconfigure(0, weight=1)
            container.rowconfigure(0, weight=1)

            canvas = tk.Canvas(container, bg="#fafafa")
            scrollbar = ttk.Scrollbar(
                container, orient="vertical", command=canvas.yview
            )
            scrollable_frame = ttk.Frame(canvas)

            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all")),
            )

            canvas.create_window(
                (0, 0),
                window=scrollable_frame,
                anchor="nw",
                width=container.winfo_width(),
            )
            canvas.configure(yscrollcommand=scrollbar.set, bg="#fafafa")

            canvas.grid(row=0, column=0, sticky="nsew")
            scrollbar.grid(row=0, column=1, sticky="ns")

            # Add mouse wheel support
            def _on_report_mousewheel(event):
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

            canvas.bind(
                "<Enter>",
                lambda e: canvas.bind_all("<MouseWheel>", _on_report_mousewheel),
            )
            canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

            # Update canvas width on container resize
            def update_canvas_width(event=None):
                if container.winfo_exists():
                    canvas_width = container.winfo_width()
                    canvas.itemconfig(
                        canvas.create_window(
                            (0, 0), window=scrollable_frame, anchor="nw"
                        ),
                        width=canvas_width,
                    )

            container.bind("<Configure>", update_canvas_width)

            # Button hover effects
            def on_enter(btn):
                btn.config(bg="#5a9bd4")

            def on_leave(btn):
                btn.config(bg="#4682b4")

            # Export to PDF button
            btn_export_pdf = tk.Button(
                btn_frame,
                text="Export to PDF",
                command=lambda: export_to_pdf(data_by_temp, table_data, ms_points_dict),
                font=("Helvetica", 10, "bold"),
                bg="#4682b4",
                fg="white",
                padx=10,
                pady=5,
                relief="flat",
                bd=2,
            )
            btn_export_pdf.pack(side=tk.RIGHT, padx=5)
            btn_export_pdf.bind("<Enter>", lambda e: on_enter(btn_export_pdf))
            btn_export_pdf.bind("<Leave>", lambda e: on_leave(btn_export_pdf))
            ToolTip(btn_export_pdf, "Export report as PDF file")

            # Export to Excel button
            btn_export_excel = tk.Button(
                btn_frame,
                text="Export to Excel",
                command=lambda: export_to_excel(
                    data_by_temp, table_data, ms_points_dict
                ),
                font=("Helvetica", 10, "bold"),
                bg="#4682b4",
                fg="white",
                padx=10,
                pady=5,
                relief="flat",
                bd=2,
            )
            btn_export_excel.pack(side=tk.RIGHT, padx=5)
            btn_export_excel.bind("<Enter>", lambda e: on_enter(btn_export_excel))
            btn_export_excel.bind("<Leave>", lambda e: on_leave(btn_export_excel))
            ToolTip(btn_export_excel, "Export report as Excel file")

            # Close button
            btn_close = tk.Button(
                btn_frame,
                text="Close",
                command=report_win.destroy,
                font=("Helvetica", 10, "bold"),
                bg="#d9534f",
                fg="white",
                padx=10,
                pady=5,
                relief="flat",
                bd=2,
            )
            btn_close.pack(side=tk.RIGHT, padx=5)
            btn_close.bind("<Enter>", lambda e: btn_close.config(bg="#e57373"))
            btn_close.bind("<Leave>", lambda e: btn_close.config(bg="#d9534f"))
            ToolTip(btn_close, "Close report window")

            # Store table data and ms_points for export
            table_data = []
            ms_points_dict = {}

            # Define export functions
            def export_to_pdf(data_by_temp, table_data, ms_points_dict):
                try:
                    num_temps = len(data_by_temp)
                    if num_temps == 0:
                        raise ValueError("No data to export.")

                    versions = set()
                    total_inflators = 0
                    temp_counts = {"RT": 0, "LT": 0, "HT": 0}
                    for temp, records in data_by_temp.items():
                        versions.update(r["version"] for r in records)
                        total_inflators += len(records)
                        temp_counts[temp] = len(records)
                    version_str = (
                        ", ".join(versions) if len(versions) > 1 else versions.pop()
                    )
                    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"report_{timestamp}.pdf"

                    # Single A4 page
                    fig = plt.figure(figsize=(8.27, 11.69))  # A4 size
                    fig.patch.set_facecolor("#fafafa")
                    fig.subplots_adjust(
                        left=0.15, right=0.85, top=0.92, bottom=0.08, hspace=0.5
                    )

                    # Title and metadata
                    ax_title = fig.add_axes([0.15, 0.92, 0.7, 0.06])
                    ax_title.axis("off")
                    ax_title.text(
                        0.5,
                        0.9,
                        "Ballistic Tests Report",
                        fontsize=12,
                        ha="center",
                        weight="bold",
                    )
                    ax_title.text(
                        0.5,
                        0.7,
                        f"Version(s): {version_str} | Total Inflators: {total_inflators}",
                        fontsize=6,
                        ha="center",
                    )
                    ax_title.text(
                        0.5,
                        0.5,
                        f"Temperatures: RT={temp_counts['RT']}, LT={temp_counts['LT']}, HT={temp_counts['HT']}",
                        fontsize=6,
                        ha="center",
                    )

                    # Data subplots (up to 3 temperatures, 2 rows each: graph + table)
                    axes = fig.subplots(
                        num_temps * 2,
                        1,
                        gridspec_kw={"height_ratios": [2.5, 1] * num_temps},
                    )
                    if num_temps == 1:
                        axes = [axes] if not isinstance(axes, np.ndarray) else axes

                    for idx, temp in enumerate(["RT", "LT", "HT"]):
                        if temp not in data_by_temp:
                            continue
                        records = data_by_temp[temp]
                        ax_graph = axes[idx * 2]
                        ax_table = axes[idx * 2 + 1]

                        versions = set(r["version"] for r in records)
                        version = (
                            ", ".join(versions)
                            if len(versions) > 1
                            else list(versions)[0]
                        )
                        total_inflators = len(records)

                        ms_points = ms_points_dict.get(temp, [])
                        ms_points_str = [str(ms) for ms in ms_points]

                        pressure_matrix = []
                        for r in records:
                            p = []
                            if r["pressures"]:
                                for ms in ms_points_str:
                                    val = r["pressures"].get(ms, np.nan)
                                    p.append(val)
                            else:
                                p = [np.nan] * len(ms_points)
                            pressure_matrix.append(p)
                        pressure_matrix = np.array(pressure_matrix, dtype=np.float64)

                        limits_max = []
                        limits_min = []
                        try:
                            with open(self.json_file, "r", encoding="utf-8") as f:
                                data_json = json.load(f)
                            sample_order = records[0]["order"]
                            limits = data_json[version][sample_order]["temperatures"][
                                temp
                            ]["limits"]
                            max_dict = limits.get("maximums", {})
                            min_dict = limits.get("minimums", {})
                            limits_max = [
                                max_dict.get(str(ms), np.nan) for ms in ms_points
                            ]
                            limits_min = [
                                min_dict.get(str(ms), np.nan) for ms in ms_points
                            ]
                        except Exception:
                            limits_max = [np.nan] * len(ms_points)
                            limits_min = [np.nan] * len(ms_points)

                        mean = np.nanmean(pressure_matrix, axis=0)

                        # Plot graph
                        ax_graph.set_facecolor("#fafafa")
                        for p in pressure_matrix:
                            ax_graph.plot(
                                ms_points,
                                p,
                                color="#444444",
                                linewidth=1,
                                alpha=0.5,
                            )
                        ax_graph.plot(
                            ms_points,
                            limits_max,
                            color="#d62728",
                            linewidth=1.5,
                            label="Maximum Limit",
                            linestyle="--",
                        )
                        ax_graph.plot(
                            ms_points,
                            limits_min,
                            color="#1f77b4",
                            linewidth=1.5,
                            label="Minimum Limit",
                            linestyle="--",
                        )
                        ax_graph.plot(
                            ms_points,
                            mean,
                            color="#2ca02c",
                            linewidth=2,
                            label="Mean",
                            linestyle="-",
                        )
                        ax_graph.set_title(
                            f"{temp} | Version: {version} | Inflators: {total_inflators}",
                            fontsize=7,
                            pad=5,
                        )
                        ax_graph.set_xlabel("Time (ms)", fontsize=5)
                        ax_graph.set_ylabel("Pressure (bar)", fontsize=5)
                        ax_graph.legend(loc="lower right", fontsize=5)
                        ax_graph.grid(
                            True, color="#cccccc", linestyle="--", linewidth=0.5
                        )
                        ax_graph.minorticks_on()
                        ax_graph.grid(
                            True,
                            which="minor",
                            color="#e0e0e0",
                            linestyle=":",
                            linewidth=0.3,
                        )
                        ax_graph.tick_params(axis="both", which="major", labelsize=5)

                        # Create table
                        def format_row(row):
                            return [f"{v:.2f}" if not np.isnan(v) else "-" for v in row]

                        table_data_rows = [
                            format_row(limits_max),
                            format_row(mean),
                            format_row(limits_min),
                        ]
                        row_labels = ["Maximum", "Mean", "Minimum"]
                        col_labels = ms_points_str

                        # Limit number of columns to avoid overflow
                        max_cols = min(len(col_labels), 12)
                        col_labels = col_labels[:max_cols]
                        table_data_rows = [row[:max_cols] for row in table_data_rows]

                        cell_colors = [
                            ["#ffcccc"] * max_cols,
                            ["#ccffcc"] * max_cols,
                            ["#cce6ff"] * max_cols,
                        ]

                        table = ax_table.table(
                            cellText=table_data_rows,
                            rowLabels=row_labels,
                            colLabels=col_labels,
                            cellColours=cell_colors,
                            cellLoc="center",
                            loc="center",
                            bbox=[0.05, 0, 0.95, 1],
                        )
                        table.auto_set_font_size(False)
                        table.set_fontsize(5)
                        table.scale(1, 1.1)
                        ax_table.axis("off")

                    # Add footer
                    fig.text(
                        0.15,
                        0.03,
                        f"Generated: {report_date}",
                        fontsize=7,
                        ha="left",
                    )

                    # Save to PDF
                    with PdfPages(filename) as pdf:
                        pdf.savefig(fig, bbox_inches="tight")
                        plt.close(fig)

                    messagebox.showinfo("Success", f"Report exported to {filename}")
                except Exception as e:
                    messagebox.showerror("Error", f"Error exporting to PDF: {str(e)}")

            def adjust_column_widths(ws):
                column_widths = {}
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value:
                            col_letter = cell.column_letter
                            current_width = column_widths.get(col_letter, 0)
                            cell_len = len(str(cell.value))
                            column_widths[col_letter] = max(current_width, cell_len)
                for col_letter, width in column_widths.items():
                    adjusted_width = min(width * 1.2, 50)
                    ws.column_dimensions[col_letter].width = adjusted_width

            def export_to_excel(data_by_temp, table_data, ms_points_dict):
                try:
                    wb = Workbook()
                    wb.remove(wb.active)
                    center_alignment = Alignment(horizontal="center", vertical="center")
                    thin_border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )
                    bold_font = Font(bold=True)
                    max_fill = PatternFill(
                        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                    )
                    mean_fill = PatternFill(
                        start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
                    )
                    min_fill = PatternFill(
                        start_color="CCE6FF", end_color="CCE6FF", fill_type="solid"
                    )

                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"report_{timestamp}.xlsx"

                    for idx, temp in enumerate(["RT", "LT", "HT"]):
                        if temp not in data_by_temp:
                            continue
                        records = data_by_temp[temp]
                        ws = wb.create_sheet(title=temp)
                        versions = set(r["version"] for r in records)
                        version = (
                            ", ".join(versions)
                            if len(versions) > 1
                            else list(versions)[0]
                        )
                        total_inflators = len(records)

                        # Metadata
                        ws.append(["Temperature", temp])
                        ws.append(["Version", version])
                        ws.append(["Total Inflators", total_inflators])
                        ws.append([])

                        # Apply styles to metadata
                        for row in ws["A1:B3"]:
                            for cell in row:
                                cell.alignment = center_alignment
                                cell.font = bold_font
                                cell.border = thin_border

                        ms_points_str = ms_points_dict.get(temp, [])
                        ws.append([""] + ms_points_str)

                        # Table headers
                        header_row = 5
                        for col_idx, ms in enumerate([""] + ms_points_str, 1):
                            cell = ws.cell(row=header_row, column=col_idx)
                            cell.value = ms
                            cell.alignment = center_alignment
                            cell.font = bold_font
                            cell.border = thin_border

                        # Table data
                        for row_idx, (row_label, row_data) in enumerate(
                            table_data[idx], start=header_row + 1
                        ):
                            ws.append([row_label] + row_data)
                            for col_idx in range(1, len(row_data) + 2):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.alignment = center_alignment
                                cell.border = thin_border
                                if row_label == "Maximum":
                                    cell.fill = max_fill
                                elif row_label == "Mean":
                                    cell.fill = mean_fill
                                elif row_label == "Minimum":
                                    cell.fill = min_fill

                        ws.append([])
                        ws.append(["Inflator Data"])
                        ws.append(["Inflator No"] + ms_points_str)

                        # Inflator data headers
                        inflator_header_row = ws.max_row
                        for col_idx, ms in enumerate(
                            ["Inflator No"] + ms_points_str, 1
                        ):
                            cell = ws.cell(row=inflator_header_row, column=col_idx)
                            cell.value = ms
                            cell.alignment = center_alignment
                            cell.font = bold_font
                            cell.border = thin_border

                        # Inflator data
                        for r in records:
                            if r["pressures"]:
                                row = [str(r["inflator_no"])]
                                for ms in ms_points_str:
                                    val = r["pressures"].get(ms, np.nan)
                                    row.append(
                                        f"{val:.2f}" if not np.isnan(val) else "-"
                                    )
                                ws.append(row)
                                for col_idx in range(1, len(row) + 1):
                                    cell = ws.cell(row=ws.max_row, column=col_idx)
                                    cell.alignment = center_alignment
                                    cell.border = thin_border

                        adjust_column_widths(ws)

                    wb.save(filename)
                    messagebox.showinfo("Success", f"Report exported to {filename}")
                except Exception as e:
                    messagebox.showerror("Error", f"Error exporting to Excel: {str(e)}")

            # For each temperature, plot graph and table
            for temp in ["RT", "LT", "HT"]:
                if temp not in data_by_temp:
                    continue
                records = data_by_temp[temp]
                versions = set(r["version"] for r in records)
                version = (
                    ", ".join(versions) if len(versions) > 1 else list(versions)[0]
                )
                total_inflators = len(records)

                temp_frame = ttk.LabelFrame(
                    scrollable_frame,
                    text=f"Temperature: {temp} | Version: {version} | Total Inflators: {total_inflators}",
                    padding=5,
                )
                temp_frame.pack(fill=tk.BOTH, expand=True, padx=0, pady=5)
                temp_frame.columnconfigure(0, weight=1)

                ms_points = set()
                for r in records:
                    if r["pressures"]:
                        ms_points.update(r["pressures"].keys())
                ms_points = sorted(int(ms) for ms in ms_points)
                ms_points_str = [str(ms) for ms in ms_points]
                ms_points_dict[temp] = ms_points_str

                if not ms_points:
                    ttk.Label(
                        temp_frame,
                        text="No pressure data available for this temperature.",
                        font=("Helvetica", 10),
                    ).pack(pady=10, fill=tk.BOTH, expand=True)
                    continue

                pressure_matrix = []
                for r in records:
                    p = []
                    if r["pressures"]:
                        for ms in ms_points_str:
                            val = r["pressures"].get(ms, np.nan)
                            p.append(val)
                    else:
                        p = [np.nan] * len(ms_points)
                    pressure_matrix.append(p)
                pressure_matrix = np.array(pressure_matrix, dtype=np.float64)

                limits_max = []
                limits_min = []
                try:
                    with open(self.json_file, "r", encoding="utf-8") as f:
                        data_json = json.load(f)
                    sample_order = records[0]["order"]
                    limits = data_json[version][sample_order]["temperatures"][temp][
                        "limits"
                    ]
                    max_dict = limits.get("maximums", {})
                    min_dict = limits.get("minimums", {})
                    limits_max = [max_dict.get(str(ms), np.nan) for ms in ms_points]
                    limits_min = [min_dict.get(str(ms), np.nan) for ms in ms_points]
                except Exception:
                    limits_max = [np.nan] * len(ms_points)
                    limits_min = [np.nan] * len(ms_points)

                mean = np.nanmean(pressure_matrix, axis=0)

                # Initial figure
                fig, ax = plt.subplots(figsize=(8, 4))
                fig.patch.set_facecolor("#fafafa")
                ax.set_facecolor("#fafafa")

                # Initial plot
                for p in pressure_matrix:
                    ax.plot(
                        ms_points,
                        p,
                        color="#444444",
                        linewidth=1,
                        alpha=0.5,
                    )
                ax.plot(
                    ms_points,
                    limits_max,
                    color="#d62728",
                    linewidth=2,
                    label="Maximum Limit",
                    linestyle="--",
                )
                ax.plot(
                    ms_points,
                    limits_min,
                    color="#1f77b4",
                    linewidth=2,
                    label="Minimum Limit",
                    linestyle="--",
                )
                ax.plot(
                    ms_points,
                    mean,
                    color="#2ca02c",
                    linewidth=2.5,
                    label="Mean",
                    linestyle="-",
                )
                ax.set_title(
                    f"Pressure Curves - Temperature {temp}", fontsize=12, pad=10
                )
                ax.set_xlabel("Time (ms)", fontsize=10)
                ax.set_ylabel("Pressure (bar)", fontsize=10)
                ax.legend(loc="lower right", fontsize=8)
                ax.grid(True, color="#cccccc", linestyle="--", linewidth=0.7)
                ax.minorticks_on()
                ax.grid(
                    True, which="minor", color="#e0e0e0", linestyle=":", linewidth=0.5
                )

                def update_graph(event=None):
                    if not temp_frame.winfo_exists():
                        return
                    # Get current dimensions of temp_frame in pixels
                    frame_width = temp_frame.winfo_width()
                    frame_height = temp_frame.winfo_height()
                    # Convert to inches (assuming 100 pixels per inch for simplicity)
                    fig_width = max(
                        4, frame_width / 100 * 0.98
                    )  # 98% of available width
                    fig_height = max(
                        2, frame_height / 100 * 0.6
                    )  # 60% of height for graph
                    fig.set_size_inches(fig_width, fig_height)
                    ax.clear()
                    ax.set_facecolor("#fafafa")
                    for p in pressure_matrix:
                        ax.plot(
                            ms_points,
                            p,
                            color="#444444",
                            linewidth=1,
                            alpha=0.5,
                        )
                    ax.plot(
                        ms_points,
                        limits_max,
                        color="#d62728",
                        linewidth=2,
                        label="Maximum Limit",
                        linestyle="--",
                    )
                    ax.plot(
                        ms_points,
                        limits_min,
                        color="#1f77b4",
                        linewidth=2,
                        label="Minimum Limit",
                        linestyle="--",
                    )
                    ax.plot(
                        ms_points,
                        mean,
                        color="#2ca02c",
                        linewidth=2.5,
                        label="Mean",
                        linestyle="-",
                    )
                    ax.set_title(
                        f"Pressure Curves - Temperature {temp}", fontsize=12, pad=10
                    )
                    ax.set_xlabel("Time (ms)", fontsize=10)
                    ax.set_ylabel("Pressure (bar)", fontsize=10)
                    ax.legend(loc="lower right", fontsize=8)
                    ax.grid(True, color="#cccccc", linestyle="--", linewidth=0.7)
                    ax.minorticks_on()
                    ax.grid(
                        True,
                        which="minor",
                        color="#e0e0e0",
                        linestyle=":",
                        linewidth=0.5,
                    )
                    canvas_fig.draw()
                    canvas_fig.get_tk_widget().update()

                temp_frame.bind("<Configure>", update_graph)

                canvas_fig = FigureCanvasTkAgg(fig, master=temp_frame)
                canvas_fig.get_tk_widget().grid(row=0, column=0, sticky="nsew", pady=5)
                temp_frame.rowconfigure(0, weight=3)
                temp_frame.rowconfigure(1, weight=1)

                # Table
                table = ttk.Treeview(
                    temp_frame, columns=ms_points_str, show="headings", height=3
                )
                for ms in ms_points_str:
                    table.heading(ms, text=ms)
                    table.column(ms, anchor="center", stretch=True)

                def format_row(row):
                    return [f"{v:.2f}" if not np.isnan(v) else "-" for v in row]

                table.insert("", "end", values=format_row(limits_max), tags=("max",))
                table.insert("", "end", values=format_row(mean), tags=("mean",))
                table.insert("", "end", values=format_row(limits_min), tags=("min",))

                table.tag_configure("max", background="#ffcccc")
                table.tag_configure("mean", background="#ccffcc")
                table.tag_configure("min", background="#cce6ff")

                style = ttk.Style()
                style.configure("Treeview", font=("Helvetica", 10))
                style.configure("Treeview.Heading", font=("Helvetica", 10, "bold"))

                def update_table_columns(event=None):
                    if not temp_frame.winfo_exists():
                        return
                    frame_width = temp_frame.winfo_width()
                    num_cols = len(ms_points_str)
                    col_width = max(
                        50, int(frame_width / num_cols * 0.98)
                    )  # 98% of width
                    for ms in ms_points_str:
                        table.column(ms, width=col_width, anchor="center")

                temp_frame.bind("<Configure>", update_table_columns)
                update_table_columns()

                table.grid(row=1, column=0, sticky="nsew", pady=5)

                table_data.append(
                    [
                        ("Maximum", format_row(limits_max)),
                        ("Mean", format_row(mean)),
                        ("Minimum", format_row(limits_min)),
                    ]
                )

            report_win.mainloop()
        except Exception as e:
            print(f"Error generating report: {str(e)}\n{traceback.format_exc()}")
            messagebox.showerror("Error", f"Error generating report: {str(e)}")


if __name__ == "__main__":
    ExcelToJsonConverter()
