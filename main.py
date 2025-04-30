import json
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from datetime import datetime
import os


class ExcelToJsonConverter:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Conversor de Ensaios Balísticos")
        self.root.geometry("1200x600")  # Tamanho inicial ajustado
        self.root.resizable(True, True)  # Permitir redimensionamento dinâmico

        # Seção para alimentar o banco de dados
        self.frame_db = tk.Frame(self.root)
        self.frame_db.pack(side=tk.LEFT, padx=10, pady=10, fill=tk.Y)

        self.label_orders = tk.Label(
            self.frame_db, text="Digite os números das ordens separados por vírgula:"
        )
        self.label_orders.pack()

        self.entry_orders = tk.Entry(self.frame_db, width=50)
        self.entry_orders.pack()

        self.btn_process = tk.Button(
            self.frame_db,
            text="Processar Ordens e Atualizar JSON",
            command=self.process_orders,
        )
        self.btn_process.pack(pady=5)

        self.status_label = tk.Label(self.frame_db, text="")
        self.status_label.pack(pady=5)

        # Seção para visualização das ordens
        self.frame_view = tk.Frame(self.root)
        self.frame_view.pack(side=tk.RIGHT, padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.label_view = tk.Label(
            self.frame_view, text="Visualizar Ordens Adicionadas com Filtros:"
        )
        self.label_view.pack()

        # Opções de filtro
        self.filter_frame = tk.Frame(self.frame_view)
        self.filter_frame.pack(pady=5)

        self.filter_version_label = tk.Label(self.filter_frame, text="Versão:")
        self.filter_version_label.pack(side=tk.LEFT, padx=5)

        self.filter_version_entry = tk.Entry(self.filter_frame, width=10)
        self.filter_version_entry.pack(side=tk.LEFT, padx=5)

        self.filter_order_label = tk.Label(self.filter_frame, text="Ordem:")
        self.filter_order_label.pack(side=tk.LEFT, padx=5)

        self.filter_order_entry = tk.Entry(self.filter_frame, width=10)
        self.filter_order_entry.pack(side=tk.LEFT, padx=5)

        self.filter_temperature_label = tk.Label(self.filter_frame, text="Temperatura:")
        self.filter_temperature_label.pack(side=tk.LEFT, padx=5)

        self.filter_temperature_entry = tk.Entry(self.filter_frame, width=10)
        self.filter_temperature_entry.pack(side=tk.LEFT, padx=5)

        self.btn_apply_filters = tk.Button(
            self.filter_frame, text="Aplicar Filtros", command=self.apply_filters
        )
        self.btn_apply_filters.pack(side=tk.LEFT, padx=5)

        # Lista de resultados com barras de rolagem integradas
        self.frame_results = tk.Frame(self.frame_view)
        self.frame_results.pack(pady=10, fill=tk.BOTH, expand=True)

        # Barra de rolagem vertical
        self.scrollbar_y = tk.Scrollbar(self.frame_results, orient=tk.VERTICAL)

        # Barra de rolagem horizontal
        self.scrollbar_x = tk.Scrollbar(self.frame_results, orient=tk.HORIZONTAL)

        # Lista de resultados (Listbox) com barras de rolagem integradas
        self.list_results = tk.Listbox(
            self.frame_results,
            width=120,
            height=25,
            yscrollcommand=self.scrollbar_y.set,
            xscrollcommand=self.scrollbar_x.set,
        )
        self.list_results.grid(row=0, column=0, sticky="nsew")

        # Configurar as barras de rolagem
        self.scrollbar_y.config(command=self.list_results.yview)
        self.scrollbar_y.grid(row=0, column=1, sticky="ns")

        self.scrollbar_x.config(command=self.list_results.xview)
        self.scrollbar_x.grid(row=1, column=0, sticky="ew")

        # Ajustar o layout (expansão automática)
        self.frame_results.grid_rowconfigure(0, weight=1)
        self.frame_results.grid_columnconfigure(0, weight=1)

        self.btn_refresh = tk.Button(
            self.frame_view, text="Atualizar Visualização", command=self.refresh_view
        )
        self.btn_refresh.pack(pady=5)

        # Arquivo JSON e pasta de Excel
        self.json_file = "Data.json"
        self.excel_folder = r"H:\TEAMS\Inflator_Lab\0_Evaluations\vi"

    def process_orders(self):
        try:
            orders_input = self.entry_orders.get().strip()
            if not orders_input:
                messagebox.showerror("Erro", "Por favor, insira os números das ordens.")
                return

            orders = [order.strip() for order in orders_input.split(",")]

            if os.path.exists(self.json_file):
                with open(self.json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
            else:
                data = {}

            files_to_process = []
            for order in orders:
                for file_name in os.listdir(self.excel_folder):
                    if file_name.startswith(order) and file_name.endswith(".xlsx"):
                        files_to_process.append(
                            os.path.join(self.excel_folder, file_name)
                        )

            if not files_to_process:
                messagebox.showerror(
                    "Erro", "Nenhum arquivo Excel encontrado para as ordens fornecidas."
                )
                return

            for file in files_to_process:
                self.process_excel(file, data)

            with open(self.json_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            self.entry_orders.delete(0, tk.END)  # Limpar campo após processamento
            self.status_label.config(text="Base de dados JSON atualizada com sucesso!")
            messagebox.showinfo(
                "Sucesso", "Arquivos Excel processados e JSON atualizado!"
            )
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar ordens: {str(e)}")

    def process_excel(self, file_path, data):
        wb = load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            if "minus" in sheet_name:
                temp_type = "LT"
            elif "RT" in sheet_name:
                temp_type = "RT"
            elif "plus" in sheet_name:
                temp_type = "HT"
            else:
                continue

            if "Datenblatt" in sheet_name:
                self.process_datenblatt(wb[sheet_name], temp_type, data)
            elif "Grafik" in sheet_name:
                self.process_grafik(wb[sheet_name], temp_type, data)

    def process_datenblatt(self, sheet, temp_type, data):
        inflator_type = self.clean_value(sheet["U1"].value)
        version = "V" + inflator_type.split("V")[-1]

        test_order = self.clean_value(sheet["J4"].value)
        production_order = self.clean_value(sheet["J3"].value)
        propellant_lot_number = self.clean_value(sheet["S3"].value)
        test_date = self.parse_date(sheet["C4"].value)

        temperature_c = self.clean_value(sheet["C10"].value)

        if version not in data:
            data[version] = {}

        if test_order not in data[version]:
            data[version][test_order] = {
                "metadados": {
                    "production_order": production_order,
                    "propellant_lot_number": propellant_lot_number,
                    "test_date": test_date,
                },
                "temperaturas": {},
            }

        if temp_type not in data[version][test_order]["temperaturas"]:
            data[version][test_order]["temperaturas"][temp_type] = {
                "temperatura_c": float(temperature_c) if temperature_c else None,
                "ensaios": [],
            }

        tests = []
        seen_tests = set()
        for row in sheet.iter_rows(min_row=10, values_only=True):
            if row[0] and str(row[0]).strip().isdigit():
                test_no = self.clean_value(row[0])
                inflator_no = self.clean_value(row[1])

                if test_no and inflator_no and test_no not in seen_tests:
                    tests.append(
                        {"test_no": int(test_no), "inflator_no": int(inflator_no)}
                    )
                    seen_tests.add(test_no)

        data[version][test_order]["temperaturas"][temp_type]["ensaios"] = tests

    def process_grafik(self, sheet, temp_type, data):
        valid_columns = []
        for col in range(3, 151):
            min_val = self.clean_value(sheet.cell(row=51, column=col).value)
            max_val = self.clean_value(sheet.cell(row=55, column=col).value)
            if min_val or max_val:
                valid_columns.append(col)

        datenblatt_inflators = [
            ensayo["inflator_no"]
            for version in data.values()
            for orders in version.values()
            for temp in orders["temperaturas"].values()
            for ensayo in temp["ensaios"]
        ]

        pressure_data = []
        for inflator_idx, inflator_no in zip(
            range(60, 60 + len(datenblatt_inflators)), datenblatt_inflators
        ):
            pressures = {}
            for col in valid_columns:
                ms = col - 2
                pressure = self.clean_value(
                    sheet.cell(row=inflator_idx, column=col).value
                )
                if pressure is not None:
                    try:
                        pressures[str(ms)] = float(pressure)
                    except ValueError:
                        continue

            if pressures:
                pressure_data.append(
                    {"inflator_no": inflator_no, "pressoes": pressures}
                )

        for version in data.values():
            for orders in version.values():
                if temp_type in orders["temperaturas"]:
                    orders["temperaturas"][temp_type]["dados_pressao"] = pressure_data

    def refresh_view(self):
        try:
            if not os.path.exists(self.json_file):
                messagebox.showerror("Erro", "Nenhum banco de dados encontrado.")
                return

            with open(self.json_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            self.list_results.delete(0, tk.END)

            for version, orders in data.items():
                for order, details in orders.items():
                    for temp_type, temp_data in details["temperaturas"].items():
                        self.list_results.insert(
                            tk.END,
                            f"Versão: {version}, Ordem: {order}, Temperatura: {temp_type}, "
                            f"Meta: {details['metadados']}, Temperatura_C: {temp_data.get('temperatura_c')}, "
                            f"Ensaios: {temp_data.get('ensaios')}",
                        )
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao atualizar visualização: {str(e)}")

    def apply_filters(self):
        try:
            if not os.path.exists(self.json_file):
                messagebox.showerror("Erro", "Nenhum banco de dados encontrado.")
                return

            with open(self.json_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            version_filter = self.filter_version_entry.get().strip()
            order_filter = self.filter_order_entry.get().strip()
            temp_filter = self.filter_temperature_entry.get().strip()

            self.list_results.delete(0, tk.END)

            for version, orders in data.items():
                if version_filter and version != version_filter:
                    continue

                for order, details in orders.items():
                    if order_filter and order != order_filter:
                        continue

                    for temp_type, temp_data in details["temperaturas"].items():
                        if temp_filter and temp_type != temp_filter:
                            continue

                        self.list_results.insert(
                            tk.END,
                            f"Versão: {version}, Ordem: {order}, Temperatura: {temp_type}, "
                            f"Meta: {details['metadados']}, Temperatura_C: {temp_data.get('temperatura_c')}, "
                            f"Ensaios: {temp_data.get('ensaios')}",
                        )
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao aplicar filtros: {str(e)}")

    def clean_value(self, value):
        if isinstance(value, str):
            value = value.strip().lstrip("'")
            if value in ("", "None", "NaN"):
                return None
            return value
        return value

    def parse_date(self, date_value):
        if isinstance(date_value, datetime):
            return date_value.strftime("%Y-%m-%d")
        return date_value

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    converter = ExcelToJsonConverter()
    converter.run()
