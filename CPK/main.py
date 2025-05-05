import os
import glob
import shutil
import json
import tkinter as tk
import numpy as np
import traceback
from datetime import datetime, timedelta
from tkcalendar import DateEntry
import config
from tkinter import ttk, messagebox
from openpyxl import load_workbook
from export_database import export_database_to_excel
from report import show_report
from utils import clean_value, parse_date
from tooltip import ToolTip


class ExcelToJsonConverter:
    def __init__(self):
        # Constantes para process_grafik
        min_limit_row = config.MIN_LIMIT_ROW
        max_limit_row = config.MAX_LIMIT_ROW
        pressure_data_start_row = config.PRESSURE_DATA_START_ROW
        min_column = config.MIN_COLUMN
        max_column = config.MAX_COLUMN

        self.root = tk.Tk()
        self.root.title("Ballistic Tests Converter")
        self.root.geometry("1200x700")
        self.root.minsize(1000, 600)
        self.root.columnconfigure(0, weight=1)
        self.root.columnconfigure(1, weight=1)
        self.root.rowconfigure(0, weight=1)

        # Left Frame (Database)
        self.frame_db = tk.Frame(self.root, relief="groove", padx=10, pady=10, bd=2)
        self.frame_db.grid(row=0, column=0, sticky="nsew")
        self.frame_db.columnconfigure(0, weight=1)
        self.frame_db.rowconfigure(5, weight=1)

        # Database Title Frame
        title_frame = tk.Frame(self.frame_db)
        title_frame.grid(row=0, column=0, sticky="ew")
        title_frame.columnconfigure(0, weight=1)

        self.label_database_title = tk.Label(
            title_frame, text="Database", font=("Helvetica", 14, "bold")
        )
        self.label_database_title.grid(row=0, column=0, sticky="w", pady=(0, 10))

        # Export Database Button
        self.btn_export_db = tk.Button(
            title_frame,
            text="Export Database",
            command=self.export_database_to_excel,
            font=("Helvetica", 12, "bold"),
            bg="#90EE90",
            fg="white",
            padx=15,
            pady=10,
        )
        self.btn_export_db.grid(row=0, column=1, sticky="e", pady=(0, 10))
        ToolTip(self.btn_export_db, "Export database to Excel")

        # Orders Input
        tk.Label(self.frame_db, text="Enter order numbers separated by commas:").grid(
            row=1, column=0, sticky="w", pady=(10, 5)
        )
        self.entry_orders = tk.Entry(self.frame_db)
        self.entry_orders.grid(row=2, column=0, sticky="ew", pady=5)

        # Process and Remove Orders Buttons Side by Side
        btns_frame = tk.Frame(self.frame_db)
        btns_frame.grid(row=3, column=0, sticky="ew", pady=5)
        btns_frame.columnconfigure((0, 1), weight=1)

        self.btn_process = tk.Button(
            btns_frame,
            text="Add Entered Orders",
            command=self.process_orders,
            font=("Helvetica", 10, "bold"),
            width=15,
        )
        self.btn_process.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        ToolTip(self.btn_process, "Process and add orders to the database")

        self.btn_remove_orders = tk.Button(
            btns_frame,
            text="Remove Entered Orders",
            command=self.remove_orders_by_input,
            font=("Helvetica", 10, "bold"),
            width=15,
        )
        self.btn_remove_orders.grid(row=0, column=1, sticky="ew", padx=(5, 0))
        ToolTip(self.btn_remove_orders, "Remove specified orders from the database")

        self.status_label = tk.Label(self.frame_db, text="", anchor="w", fg="green")
        self.status_label.grid(row=4, column=0, sticky="ew", pady=(5, 10))

        # Orders Manager
        self.frame_orders_manager = tk.Frame(
            self.frame_db, relief="groove", bd=2, padx=10, pady=10
        )
        self.frame_orders_manager.grid(row=5, column=0, sticky="nsew")
        self.frame_orders_manager.columnconfigure(0, weight=1)
        self.frame_orders_manager.rowconfigure(3, weight=1)  # Canvas row expands

        # Initialize Pagination Variables
        self.current_page = 1
        self.orders_per_page = 10
        self.total_pages = 1

        # Orders Manager Title
        title_orders_frame = tk.Frame(self.frame_orders_manager)
        title_orders_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        title_orders_frame.columnconfigure(0, weight=1)
        tk.Label(
            title_orders_frame, text="Orders Manager", font=("Helvetica", 12, "bold")
        ).grid(row=0, column=0, sticky="w")

        # Filters Frame
        filters_frame = tk.Frame(
            self.frame_orders_manager,
            relief="groove",
            bd=2,
            bg="#f0f0f0",
            padx=10,
            pady=10,
        )
        filters_frame.grid(row=1, column=0, sticky="ew", pady=(0, 5))
        filters_frame.columnconfigure((0, 1, 2, 3), weight=1)

        # Version Filter
        tk.Label(filters_frame, text="Version:", bg="#f0f0f0").grid(
            row=0, column=0, sticky="w", padx=(0, 5)
        )
        self.version_var = tk.StringVar()
        self.version_combobox = ttk.Combobox(
            filters_frame,
            textvariable=self.version_var,
            state="readonly",
            width=10,
        )
        self.version_combobox.grid(row=0, column=1, sticky="w")
        self.version_combobox.bind("<<ComboboxSelected>>", self.on_version_filter)
        self.version_combobox["values"] = []
        self.version_combobox.set("All")
        ToolTip(self.version_combobox, "Filter orders by version")

        # Date Range Filter
        tk.Label(filters_frame, text="Date Range:", bg="#f0f0f0").grid(
            row=0, column=2, sticky="w", padx=(10, 5)
        )
        self.date_range_var = tk.StringVar()
        self.date_range_combobox = ttk.Combobox(
            filters_frame,
            textvariable=self.date_range_var,
            state="readonly",
            values=["All", "Last 30 Days", "Last 60 Days", "Last 90 Days", "Custom"],
            width=12,
        )
        self.date_range_combobox.grid(row=0, column=3, sticky="w")
        self.date_range_combobox.set("All")
        self.date_range_combobox.bind("<<ComboboxSelected>>", self.on_date_range_filter)
        ToolTip(self.date_range_combobox, "Filter orders by date range")

        # Custom Date Inputs
        self.custom_dates_frame = tk.Frame(filters_frame, bg="#f0f0f0")
        self.custom_dates_frame.grid(
            row=1, column=0, columnspan=4, sticky="w", pady=(5, 0)
        )
        tk.Label(self.custom_dates_frame, text="Start:", bg="#f0f0f0").grid(
            row=0, column=0, sticky="w", padx=(0, 5)
        )
        self.start_date_entry = tk.Entry(self.custom_dates_frame, width=12)
        self.start_date_entry.grid(row=0, column=1, sticky="w")
        self.start_date_btn = tk.Button(
            self.custom_dates_frame,
            text="📅",
            command=lambda: self.pick_date(self.start_date_entry),
            width=2,
        )
        self.start_date_btn.grid(row=0, column=2, sticky="w", padx=2)
        tk.Label(self.custom_dates_frame, text="End:", bg="#f0f0f0").grid(
            row=0, column=3, sticky="w", padx=(10, 5)
        )
        self.end_date_entry = tk.Entry(self.custom_dates_frame, width=12)
        self.end_date_entry.grid(row=0, column=4, sticky="w")
        self.end_date_btn = tk.Button(
            self.custom_dates_frame,
            text="📅",
            command=lambda: self.pick_date(self.end_date_entry),
            width=2,
        )
        self.end_date_btn.grid(row=0, column=5, sticky="w", padx=2)
        self.custom_dates_frame.grid_remove()  # Hidden by default

        # Pagination Frame
        pagination_frame = tk.Frame(self.frame_orders_manager)
        pagination_frame.grid(row=2, column=0, sticky="ew", pady=5)
        pagination_frame.columnconfigure(2, weight=1)

        tk.Label(pagination_frame, text="Items per page:").grid(
            row=0, column=0, sticky="w", padx=(0, 5)
        )
        self.page_selector = ttk.Combobox(
            pagination_frame,
            values=[5, 10, 15, 20, 25, 30],
            width=5,
            state="readonly",
        )
        self.page_selector.set(self.orders_per_page)
        self.page_selector.grid(row=0, column=1, sticky="w")
        self.page_selector.bind("<<ComboboxSelected>>", self.on_items_per_page_changed)
        ToolTip(self.page_selector, "Select number of orders per page")

        self.select_all_var = tk.BooleanVar()
        self.select_all_chk = tk.Checkbutton(
            pagination_frame,
            text="Select All",
            variable=self.select_all_var,
            command=self.toggle_select_all,
        )
        self.select_all_chk.grid(row=0, column=2, sticky="w", padx=10)
        ToolTip(self.select_all_chk, "Select all orders on current page")

        self.nav_frame = tk.Frame(pagination_frame)
        self.nav_frame.grid(row=0, column=3, sticky="e")
        self.prev_btn = tk.Button(
            self.nav_frame,
            text="< Prev",
            command=lambda: self.change_page(-1),
            width=8,
            font=("Helvetica", 10),
        )
        self.prev_btn.pack(side="left", padx=2)
        self.page_info = tk.Label(
            self.nav_frame, text="Page 1/1", width=10, anchor="center"
        )
        self.page_info.pack(side="left", padx=5)
        self.next_btn = tk.Button(
            self.nav_frame,
            text="Next >",
            command=lambda: self.change_page(1),
            width=8,
            font=("Helvetica", 10),
        )
        self.next_btn.pack(side="left", padx=2)

        # Canvas and Scrollbar for Orders
        self.orders_canvas = tk.Canvas(self.frame_orders_manager)
        self.orders_canvas.grid(row=3, column=0, sticky="nsew", padx=5, pady=5)
        self.orders_scrollbar = tk.Scrollbar(
            self.frame_orders_manager,
            orient=tk.VERTICAL,
            command=self.orders_canvas.yview,
        )
        self.orders_scrollbar.grid(row=3, column=1, sticky="ns", pady=5)
        self.orders_canvas.configure(yscrollcommand=self.orders_scrollbar.set)

        self.orders_inner_frame = tk.Frame(self.orders_canvas)
        self.orders_canvas.create_window(
            (0, 0), window=self.orders_inner_frame, anchor="nw"
        )

        self.orders_inner_frame.bind(
            "<Configure>",
            lambda e: self.orders_canvas.configure(
                scrollregion=self.orders_canvas.bbox("all")
            ),
        )
        self.orders_canvas.yview_moveto(0)
        self.orders_canvas.bind(
            "<Enter>",
            lambda e: self.orders_canvas.bind_all("<MouseWheel>", self._on_mousewheel),
        )
        self.orders_canvas.bind(
            "<Leave>", lambda e: self.orders_canvas.unbind_all("<MouseWheel>")
        )

        self.order_vars = {}
        self.order_checkbuttons = {}

        # Action Buttons Frame
        action_btn_frame = tk.Frame(
            self.frame_orders_manager, relief="groove", bd=2, padx=10, pady=10
        )
        action_btn_frame.grid(row=4, column=0, sticky="ew", pady=5)
        action_btn_frame.columnconfigure((0, 1, 2), weight=1)

        self.btn_send_workplace = tk.Button(
            action_btn_frame,
            text="Send to Workplace",
            command=self.send_to_workplace,
            font=("Helvetica", 10, "bold"),
            bg="#90EE90",
            width=15,
        )
        self.btn_send_workplace.grid(row=0, column=0, sticky="ew", padx=5)
        ToolTip(self.btn_send_workplace, "Send selected orders to workplace")

        self.btn_remove_workplace_orders = tk.Button(
            action_btn_frame,
            text="Remove Tests",
            command=self.remove_workplace_orders_selected,
            font=("Helvetica", 10, "bold"),
            bg="#FF6B6B",
            width=15,
        )
        self.btn_remove_workplace_orders.grid(row=0, column=1, sticky="ew", padx=5)
        ToolTip(
            self.btn_remove_workplace_orders, "Remove selected orders from workplace"
        )

        self.btn_clear_workplace = tk.Button(
            action_btn_frame,
            text="Clear Workplace",
            command=self.clear_workplace,
            font=("Helvetica", 10, "bold"),
            bg="#D3D3D3",
            width=15,
        )
        self.btn_clear_workplace.grid(row=0, column=2, sticky="ew", padx=5)
        ToolTip(self.btn_clear_workplace, "Clear all tests from workplace")

        # Workplace (right)
        self.workplace_frame = tk.Frame(
            self.root, relief="groove", bd=2, padx=10, pady=10
        )
        self.workplace_frame.grid(row=0, column=1, sticky="nsew")
        self.workplace_frame.columnconfigure(0, weight=1)
        self.workplace_frame.rowconfigure(2, weight=1)

        # Frame for Report and Close buttons aligned to the right at the top of the Workplace
        btn_report_frame = tk.Frame(self.workplace_frame)
        btn_report_frame.grid(row=0, column=0, sticky="ew")
        btn_report_frame.columnconfigure(0, weight=1)

        self.btn_report = tk.Button(
            btn_report_frame,
            text="Report",
            command=self.show_report,
            font=("Helvetica", 12, "bold"),
            bg="#4682b4",
            fg="white",
            padx=15,
            pady=10,
        )
        self.btn_report.grid(row=0, column=1, sticky="e", padx=(0, 5))
        ToolTip(self.btn_report, "Generate a report of workplace tests")

        self.btn_close_main = tk.Button(
            btn_report_frame,
            text="Close",
            command=self.close_application,
            font=("Helvetica", 12, "bold"),
            bg="#d9534f",
            fg="white",
            padx=15,
            pady=10,
        )
        self.btn_close_main.grid(row=0, column=2, sticky="e")
        self.btn_close_main.bind(
            "<Enter>", lambda e: self.btn_close_main.config(bg="#e57373")
        )
        self.btn_close_main.bind(
            "<Leave>", lambda e: self.btn_close_main.config(bg="#d9534f")
        )
        ToolTip(self.btn_close_main, "Close the application")

        self.workplace_title = tk.Label(
            self.workplace_frame, text="Workplace", font=("Helvetica", 14, "bold")
        )
        self.workplace_title.grid(row=0, column=0, sticky="w", pady=(0, 10))

        # Temperature and Limiter Filters
        filter_temp_frame = tk.Frame(self.workplace_frame)
        filter_temp_frame.grid(row=1, column=0, sticky="w", pady=5)

        tk.Label(filter_temp_frame, text="Filter Temperature:").pack(
            side="left", padx=(0, 5)
        )
        self.filter_temperature_var = tk.StringVar()
        self.filter_temperature_combobox = ttk.Combobox(
            filter_temp_frame,
            textvariable=self.filter_temperature_var,
            state="readonly",
            values=["All", "RT", "LT", "HT"],
            width=10,
        )
        self.filter_temperature_combobox.pack(side="left")
        self.filter_temperature_combobox.set("All")

        tk.Label(filter_temp_frame, text="Limiter:").pack(side="left", padx=(10, 5))
        self.limit_var = tk.StringVar()
        limit_values = ["All"] + [str(i) for i in range(0, 201, 1)]
        self.limit_combobox = ttk.Combobox(
            filter_temp_frame,
            textvariable=self.limit_var,
            state="readonly",
            values=limit_values,
            width=8,
        )
        self.limit_combobox.pack(side="left")
        self.limit_combobox.set("All")

        self.btn_apply_filters = tk.Button(
            filter_temp_frame, text="Apply Filters", command=self.apply_filters
        )
        self.btn_apply_filters.pack(side="left", padx=10)
        ToolTip(self.btn_apply_filters, "Apply temperature and limiter filters")

        # Results List with Scrollbars
        self.frame_results = tk.Frame(self.workplace_frame)
        self.frame_results.grid(row=2, column=0, sticky="nsew", pady=5)
        self.frame_results.columnconfigure(0, weight=1)
        self.frame_results.rowconfigure(0, weight=1)

        separator = tk.Frame(self.workplace_frame, height=2, bd=1, relief="sunken")
        separator.grid(row=3, column=0, sticky="ew", pady=(5, 5))

        self.counter_frame = tk.Frame(self.workplace_frame, pady=8)
        self.counter_frame.grid(row=4, column=0, sticky="ew")
        self.counter_frame.columnconfigure((0, 1, 2, 3), weight=1)

        self.label_rt = tk.Label(
            self.counter_frame,
            text="RT: 0",
            font=("Helvetica", 11, "bold"),
            fg="#008800",
        )
        self.label_rt.grid(row=0, column=0, sticky="ew", padx=8)

        self.label_lt = tk.Label(
            self.counter_frame,
            text="LT: 0",
            font=("Helvetica", 11, "bold"),
            fg="#0055cc",
        )
        self.label_lt.grid(row=0, column=1, sticky="ew", padx=8)

        self.label_ht = tk.Label(
            self.counter_frame,
            text="HT: 0",
            font=("Helvetica", 11, "bold"),
            fg="#cc5500",
        )
        self.label_ht.grid(row=0, column=2, sticky="ew", padx=8)

        self.label_total = tk.Label(
            self.counter_frame,
            text="Total: 0",
            font=("Helvetica", 11, "bold"),
            fg="#222222",
        )
        self.label_total.grid(row=0, column=3, sticky="ew", padx=8)

        self.scrollbar_y = tk.Scrollbar(self.frame_results, orient=tk.VERTICAL)
        self.scrollbar_x = tk.Scrollbar(self.frame_results, orient=tk.HORIZONTAL)

        self.list_results = tk.Listbox(
            self.frame_results,
            width=100,
            height=25,
            yscrollcommand=self.scrollbar_y.set,
            xscrollcommand=self.scrollbar_x.set,
        )
        self.list_results.grid(row=0, column=0, sticky="nsew")

        self.scrollbar_y.config(command=self.list_results.yview)
        self.scrollbar_y.grid(row=0, column=1, sticky="ns")

        self.scrollbar_x.config(command=self.list_results.xview)
        self.scrollbar_x.grid(row=1, column=0, sticky="ew")

        # Initializations
        self.json_file = config.JSON_FILE
        self.excel_folder = config.EXCEL_FOLDER
        self.create_daily_backup()
        self.workplace_data = []
        self.filtered_workplace_data = None

        self.update_orders_list()
        self.root.mainloop()

    def export_database_to_excel(self):
        export_database_to_excel(self)

    def show_report(self):
        show_report(self)

    def close_application(self):
        try:
            self.root.destroy()
            import sys

            sys.exit(0)
        except Exception as e:
            print(f"Error closing application: {str(e)}\n{traceback.format_exc()}")
            raise

    def _on_mousewheel(self, event):
        self.orders_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def pick_date(self, entry):
        """Open a DateEntry calendar to pick a date and insert it into the entry."""
        popup = tk.Toplevel(self.root)
        popup.title("Select Date")
        popup.geometry("250x250")
        popup.resizable(False, False)
        date_picker = DateEntry(popup, date_pattern="yyyy-mm-dd", width=12)
        date_picker.pack(pady=10)
        tk.Button(
            popup,
            text="Confirm",
            command=lambda: [
                entry.delete(0, tk.END),
                entry.insert(0, date_picker.get()),
                popup.destroy(),
                self.update_orders_list(),
            ],
        ).pack(pady=5)
        tk.Button(popup, text="Cancel", command=popup.destroy).pack(pady=5)

    def on_date_range_filter(self, event=None):
        """Handle date range filter selection and show/hide custom date inputs."""
        date_range = self.date_range_var.get()
        if date_range == "Custom":
            self.custom_dates_frame.grid()
        else:
            self.custom_dates_frame.grid_remove()
        self.current_page = 1
        for var in self.order_vars.values():
            var.set(False)
        self.select_all_var.set(False)
        self.update_orders_list()

    def create_daily_backup(self):
        try:
            if not os.path.exists(self.json_file):
                print(f"Backup skipped: {self.json_file} not found.")
                return

            base_dir = os.path.dirname(self.json_file) or "."
            backup_dir = os.path.join(base_dir, "Backup")

            if not os.path.exists(backup_dir):
                os.makedirs(backup_dir)
                print(f"Created Backup directory: {backup_dir}")

            current_date = datetime.now().strftime("%Y%m%d")

            backup_pattern = os.path.join(backup_dir, f"Data_{current_date}_*.json")
            existing_backups = glob.glob(backup_pattern)

            if existing_backups:
                print(
                    f"Backup already exists for {current_date}: {existing_backups[0]}"
                )
                return

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"Data_{timestamp}.json"
            backup_path = os.path.join(backup_dir, backup_filename)

            shutil.copy2(self.json_file, backup_path)
            print(f"Backup created: {backup_path}")

        except Exception as e:
            error_message = f"Error creating backup: {str(e)}"
            print(error_message)

    def process_orders(self):
        try:
            orders_input = self.entry_orders.get().strip()
            if not orders_input:
                messagebox.showerror("Error", "Please enter the order numbers.")
                return

            orders = [order.strip() for order in orders_input.split(",")]

            if os.path.exists(self.json_file):
                with open(self.json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
            else:
                data = {}

            files_to_process = []
            for order in orders:
                for file_name in os.listdir(self.excel_folder):
                    if file_name.startswith(order) and file_name.endswith(".xlsx"):
                        files_to_process.append(
                            os.path.join(self.excel_folder, file_name)
                        )

            if not files_to_process:
                messagebox.showerror(
                    "Error", "No Excel files found for the provided orders."
                )
                return

            for file in files_to_process:
                self.process_excel(file, data)

            with open(self.json_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            self.entry_orders.delete(0, tk.END)
            self.status_label.config(text="JSON database updated successfully!")
            messagebox.showinfo("Success", "Excel files processed and JSON updated!")
            self.update_orders_list()
        except Exception as e:
            messagebox.showerror("Error", f"Error processing orders: {str(e)}")

    def process_excel(self, file_path, data):
        wb = load_workbook(file_path, data_only=True)
        current_version = None
        current_order = None
        for sheet_name in wb.sheetnames:
            if "minus" in sheet_name.lower():
                temp_type = "LT"
            elif "rt" in sheet_name.lower():
                temp_type = "RT"
            elif "plus" in sheet_name.lower():
                temp_type = "HT"
            else:
                continue

            if "datenblatt" in sheet_name.lower():
                current_version, current_order = self.process_datenblatt(
                    wb[sheet_name], temp_type, data
                )
            elif "grafik" in sheet_name.lower() and current_version and current_order:
                self.process_grafik(
                    wb[sheet_name], temp_type, data, current_version, current_order
                )

    def process_datenblatt(self, sheet, temp_type, data):
        inflator_type = clean_value(sheet["U1"].value)
        version = "V" + inflator_type.split("V")[-1]

        test_order = clean_value(sheet["J4"].value)
        production_order = clean_value(sheet["J3"].value)
        propellant_lot_number = clean_value(sheet["S3"].value)
        test_date = parse_date(sheet["C4"].value)
        temperature_c = clean_value(sheet["C10"].value)

        if version not in data:
            data[version] = {}

        if test_order not in data[version]:
            data[version][test_order] = {
                "metadata": {
                    "production_order": production_order,
                    "propellant_lot_number": propellant_lot_number,
                    "test_date": test_date,
                },
                "temperatures": {},
            }

        if temp_type not in data[version][test_order]["temperatures"]:
            data[version][test_order]["temperatures"][temp_type] = {
                "temperature_c": float(temperature_c) if temperature_c else None,
                "tests": [],
            }

        tests = []
        seen_tests = set()
        for row in sheet.iter_rows(min_row=10, values_only=True):
            if row[0] and str(row[0]).strip().isdigit():
                test_no = clean_value(row[0])
                inflator_no = clean_value(row[1])
                if test_no and inflator_no and test_no not in seen_tests:
                    tests.append(
                        {"test_no": int(test_no), "inflator_no": int(inflator_no)}
                    )
                    seen_tests.add(test_no)

        data[version][test_order]["temperatures"][temp_type]["tests"] = tests
        return version, test_order

    def process_grafik(self, sheet, temp_type, data, current_version, current_order):
        valid_columns = []
        limits = {"maximums": {}, "minimums": {}}
        for col in range(self.MIN_COLUMN, self.MAX_COLUMN):
            min_val = clean_value(sheet.cell(row=self.MIN_LIMIT_ROW, column=col).value)
            max_val = clean_value(sheet.cell(row=self.MAX_LIMIT_ROW, column=col).value)
            if min_val or max_val:
                valid_columns.append(col)
                ms = col - 2
                if min_val is not None:
                    try:
                        limits["minimums"][str(ms)] = float(min_val)
                    except ValueError:
                        pass
                if max_val is not None:
                    try:
                        limits["maximums"][str(ms)] = float(max_val)
                    except ValueError:
                        pass

        inflator_nos = [
            test["inflator_no"]
            for test in data[current_version][current_order]["temperatures"][temp_type][
                "tests"
            ]
        ]

        pressure_data = []
        blank_line_count = 0
        row_idx = self.PRESSURE_DATA_START_ROW

        for inflator_no in inflator_nos:
            is_blank = True
            pressures = {}
            for col in valid_columns:
                pressure = clean_value(sheet.cell(row=row_idx, column=col).value)
                if pressure is not None:
                    try:
                        pressures[str(col - 2)] = float(pressure)
                        is_blank = False
                    except ValueError:
                        continue

            if is_blank:
                blank_line_count += 1
                if blank_line_count >= 2:
                    break
            else:
                blank_line_count = 0
                if pressures:
                    pressure_data.append(
                        {"inflator_no": inflator_no, "pressures": pressures}
                    )

            row_idx += 1

        data[current_version][current_order]["temperatures"][temp_type][
            "pressure_data"
        ] = pressure_data
        data[current_version][current_order]["temperatures"][temp_type][
            "limits"
        ] = limits

    def update_orders_list(self):
        for widget in self.orders_inner_frame.winfo_children():
            widget.destroy()
        self.order_checkbuttons.clear()

        if not os.path.exists(self.json_file):
            return

        try:
            with open(self.json_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            versions = sorted(data.keys())
            self.version_combobox["values"] = ["All"] + versions
            current = self.version_combobox.get()
            if current not in self.version_combobox["values"]:
                self.version_combobox.set("All")
            selected_version = self.version_combobox.get()

            # Get date range filter
            date_range = self.date_range_var.get()
            today = datetime.now().date()
            start_date = None
            end_date = today

            if date_range == "Last 30 Days":
                start_date = today - timedelta(days=30)
            elif date_range == "Last 60 Days":
                start_date = today - timedelta(days=60)
            elif date_range == "Last 90 Days":
                start_date = today - timedelta(days=90)
            elif date_range == "Custom":
                try:
                    start_date_str = self.start_date_entry.get()
                    end_date_str = self.end_date_entry.get()
                    if start_date_str and end_date_str:
                        start_date = datetime.strptime(
                            start_date_str, "%Y-%m-%d"
                        ).date()
                        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                        if start_date > end_date:
                            messagebox.showerror(
                                "Error", "Start date cannot be after end date."
                            )
                            return
                    else:
                        start_date = None  # Treat as "All" if custom dates are empty
                except ValueError:
                    messagebox.showerror(
                        "Error", "Invalid date format. Use YYYY-MM-DD."
                    )
                    return

            orders_list = []
            for version, orders in data.items():
                if (
                    selected_version
                    and selected_version.lower() != "all"
                    and version != selected_version
                ):
                    continue
                for order, details in orders.items():
                    test_date = details["metadata"].get("test_date", "0000-00-00")
                    if start_date:
                        try:
                            test_date_obj = datetime.strptime(
                                test_date, "%Y-%m-%d"
                            ).date()
                            if not (start_date <= test_date_obj <= end_date):
                                continue
                        except ValueError:
                            continue  # Skip invalid dates
                    orders_list.append((version, order, test_date))

            def parse_date_safe(date_str):
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except Exception:
                    return datetime.min

            orders_list.sort(key=lambda x: parse_date_safe(x[2]), reverse=True)

            total_orders = len(orders_list)
            self.total_pages = max(
                1, (total_orders + self.orders_per_page - 1) // self.orders_per_page
            )
            if not hasattr(self, "current_page") or self.current_page < 1:
                self.current_page = 1
            if self.current_page > self.total_pages:
                self.current_page = self.total_pages

            start_idx = (self.current_page - 1) * self.orders_per_page
            end_idx = start_idx + self.orders_per_page
            paginated_orders = orders_list[start_idx:end_idx]

            for idx, (version, order, test_date) in enumerate(
                paginated_orders, start=start_idx + 1
            ):
                key = (version, order)
                if key not in self.order_vars:
                    self.order_vars[key] = tk.BooleanVar()
                var = self.order_vars[key]
                display_text = (
                    f"{idx}. Version: {version}, Order: {order}, Date: {test_date}"
                )

                row_frame = tk.Frame(self.orders_inner_frame)
                row_frame.grid(
                    row=idx - start_idx, column=0, sticky="w", padx=5, pady=2
                )

                chk = tk.Checkbutton(
                    row_frame,
                    text=display_text,
                    variable=var,
                    anchor="w",
                    width=40,
                )
                chk.pack(side=tk.LEFT)

                btn_view = tk.Button(
                    row_frame,
                    text="     👁️",
                    width=3,
                    command=lambda v=version, o=order: self.show_metadata_popup(v, o),
                )
                btn_view.pack(side=tk.LEFT, padx=(10, 0))

                self.order_checkbuttons[key] = chk

            self.orders_canvas.configure(scrollregion=self.orders_canvas.bbox("all"))
            self.orders_canvas.yview_moveto(0)

            self.page_info.config(text=f"Page {self.current_page}/{self.total_pages}")

            self.prev_btn.config(
                state=tk.NORMAL if self.current_page > 1 else tk.DISABLED
            )
            self.next_btn.config(
                state=tk.NORMAL if self.current_page < self.total_pages else tk.DISABLED
            )

            current_page_orders = [
                (version, order) for version, order, _ in paginated_orders
            ]
            all_selected = all(
                self.order_vars.get(key, tk.BooleanVar(value=False)).get()
                for key in current_page_orders
            )
            self.select_all_var.set(all_selected)

        except Exception as e:
            messagebox.showerror("Error", f"Error loading orders: {str(e)}")

    def show_metadata_popup(self, version, order):
        if not os.path.exists(self.json_file):
            messagebox.showerror("Error", "Database not found.")
            return

        with open(self.json_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        if version not in data or order not in data[version]:
            messagebox.showerror("Error", "Order not found in the database.")
            return

        metadata = data[version][order].get("metadata", {})
        temperatures = data[version][order].get("temperatures", {})

        info = ""
        for k, v in metadata.items():
            info += f"{k}: {v}\n"
        info += "\nTemperatures (°C):\n"
        for tipo, tdata in temperatures.items():
            info += f"  {tipo}: {tdata.get('temperature_c', 'N/A')}\n"

        popup = tk.Toplevel(self.root)
        popup.title(f"Metadata - {order}")
        popup.geometry("350x250")
        popup.resizable(False, False)
        tk.Label(
            popup,
            text=f"Metadata for Order {order} ({version})",
            font=("Arial", 11, "bold"),
        ).pack(pady=8)
        text = tk.Text(popup, width=40, height=10, wrap="word")
        text.insert("1.0", info)
        text.config(state="disabled")
        text.pack(padx=8, pady=8)
        tk.Button(popup, text="Close", command=popup.destroy).pack(pady=5)

    def update_items_per_page(self, event=None):
        try:
            self.orders_per_page = int(self.page_selector.get())
            self.current_page = 1
            self.update_orders_list()
        except Exception as e:
            messagebox.showerror("Error", f"Error updating items per page: {str(e)}")

    def on_items_per_page_changed(self, event=None):
        try:
            self.orders_per_page = int(self.page_selector.get())
        except:
            self.orders_per_page = 10
        self.current_page = 1
        self.update_orders_list()

    def change_page(self, delta):
        new_page = self.current_page + delta
        if 1 <= new_page <= self.total_pages:
            self.current_page = new_page
            self.update_orders_list()

    def toggle_select_all(self):
        state = self.select_all_var.get()
        start_idx = (self.current_page - 1) * self.orders_per_page
        end_idx = start_idx + self.orders_per_page

        try:
            with open(self.json_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            orders_list = []
            selected_version = self.version_combobox.get()
            for version, orders in data.items():
                if (
                    selected_version
                    and selected_version.lower() != "all"
                    and version != selected_version
                ):
                    continue
                for order, details in orders.items():
                    test_date = details["metadata"].get("test_date", "0000-00-00")
                    orders_list.append((version, order, test_date))

            def parse_date_safe(date_str):
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except Exception:
                    return datetime.min

            orders_list.sort(key=lambda x: parse_date_safe(x[2]), reverse=True)
            paginated_orders = orders_list[start_idx:end_idx]

            for version, order, _ in paginated_orders:
                key = (version, order)
                if key in self.order_vars:
                    self.order_vars[key].set(state)

        except Exception as e:
            messagebox.showerror("Error", f"Error toggling select all: {str(e)}")

    def on_version_filter(self, event=None):
        self.current_page = 1
        for var in self.order_vars.values():
            var.set(False)
        self.select_all_var.set(False)
        self.update_orders_list()

    def remove_orders_by_input(self):
        try:
            orders_input = self.entry_orders.get().strip()
            if not orders_input:
                messagebox.showerror("Error", "Please enter the order numbers.")
                return

            orders_to_remove = [order.strip() for order in orders_input.split(",")]

            if not os.path.exists(self.json_file):
                messagebox.showerror("Error", "No database found.")
                return

            with open(self.json_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            removed = []
            for version in list(data.keys()):
                for order in list(data[version].keys()):
                    if order in orders_to_remove:
                        del data[version][order]
                        removed.append(order)
                        if not data[version]:
                            del data[version]

            with open(self.json_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            if removed:
                msg = f"Orders removed successfully:\n{', '.join(removed)}"
                self.status_label.config(text=msg)
                messagebox.showinfo("Success", msg)
            else:
                messagebox.showwarning("Warning", "No matching orders found.")

            self.entry_orders.delete(0, tk.END)
            self.update_orders_list()

        except Exception as e:
            messagebox.showerror("Error", f"Error removing orders: {str(e)}")

    def send_to_workplace(self):
        selected_orders = [
            (version, order)
            for (version, order), var in self.order_vars.items()
            if var.get()
        ]
        if not selected_orders:
            messagebox.showwarning(
                "Warning", "No orders selected to send to workplace."
            )
            return

        try:
            with open(self.json_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            versions = set(version for version, order in selected_orders)
            if len(versions) > 1:
                messagebox.showerror("Error", "Select tests from only one version.")
                return

            if (
                self.workplace_data
                and self.workplace_data[0]["version"] not in versions
            ):
                messagebox.showerror(
                    "Error",
                    "Workplace already contains tests from another version. Clear workplace first.",
                )
                return

            existing_keys = set(
                (
                    reg["test_no"],
                    reg["inflator_no"],
                    reg["type"],
                    reg["version"],
                    reg["order"],
                )
                for reg in self.workplace_data
            )

            new_workplace_data = []
            duplicates_skipped = 0

            for version, order in selected_orders:
                if version in data and order in data[version]:
                    details = data[version][order]
                    metadata = details.get("metadata", {})
                    test_date = metadata.get("test_date", "0000-00-00")
                    temperatures = details.get("temperatures", {})
                    for temp_type in ["RT", "LT", "HT"]:
                        if temp_type not in temperatures:
                            continue
                        temp_data = temperatures[temp_type]
                        temperature_c = temp_data.get("temperature_c", "N/A")
                        tests = temp_data.get("tests", [])
                        pressure_data = temp_data.get("pressure_data", [])
                        pressure_map = {
                            item["inflator_no"]: item["pressures"]
                            for item in pressure_data
                        }
                        for test in tests:
                            test_no = test.get("test_no", "N/A")
                            inflator_no = test.get("inflator_no", "N/A")
                            key = (test_no, inflator_no, temp_type, version, order)
                            if key in existing_keys:
                                duplicates_skipped += 1
                                continue
                            new_workplace_data.append(
                                {
                                    "test_no": test_no,
                                    "inflator_no": inflator_no,
                                    "temperature_c": temperature_c,
                                    "type": temp_type,
                                    "version": version,
                                    "order": order,
                                    "test_date": test_date,
                                    "pressures": pressure_map.get(inflator_no, {}),
                                }
                            )
                            existing_keys.add(key)

            self.workplace_data.extend(new_workplace_data)

            def parse_date_safe(date_str):
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except:
                    return datetime(1900, 1, 1)

            self.workplace_data.sort(
                key=lambda x: parse_date_safe(x.get("test_date", "1900-01-01")),
                reverse=True,
            )

            self.update_workplace_display()
            self.update_workplace_counters()

            msg = f"Tests sent to workplace successfully! Added {len(new_workplace_data)} records."
            if duplicates_skipped > 0:
                msg += f"\n{duplicates_skipped} duplicate test(s) were ignored."
            messagebox.showinfo("Success", msg)

        except Exception as e:
            messagebox.showerror("Error", f"Error sending tests to workplace: {str(e)}")

    def update_workplace_display(self):
        self.list_results.delete(0, tk.END)
        header = "Test | Inflator | Temperature | Type | Version | Order | Date"
        self.list_results.insert(tk.END, header)
        self.list_results.insert(tk.END, "-" * len(header))

        for reg in self.workplace_data:
            line = f"{reg['test_no']} | {reg['inflator_no']} | {reg['temperature_c']}°C | {reg['type']} | {reg['version']} | {reg['order']} | {reg['test_date']}"
            if reg["pressures"]:
                line += " | Pressure data available"
            else:
                line += " | No pressure data"
            self.list_results.insert(tk.END, line)
        self.update_workplace_counters()

    def apply_filters(self):
        temp_filter = self.filter_temperature_var.get()
        limit_filter = self.limit_var.get()

        self.list_results.delete(0, tk.END)
        header = "Test | Inflator | Temperature | Type | Version | Order | Date"
        self.list_results.insert(tk.END, header)
        self.list_results.insert(tk.END, "-" * len(header))

        versions = {reg["version"] for reg in self.workplace_data}
        if len(versions) > 1:
            messagebox.showerror(
                "Error",
                "Workplace contains mixed versions! Clear before applying filters.",
            )
            return

        if temp_filter == "All" and limit_filter == "All":
            self.filtered_workplace_data = None
            for reg in self.workplace_data:
                line = f"{reg['test_no']} | {reg['inflator_no']} | {reg['temperature_c']}°C | {reg['type']} | {reg['version']} | {reg['order']} | {reg['test_date']}"
                if reg["pressures"]:
                    line += " | Pressure data available"
                else:
                    line += " | No pressure data"
                self.list_results.insert(tk.END, line)
            self.update_workplace_counters()
            return

        filtered_data = []
        for reg in self.workplace_data:
            if temp_filter != "All" and reg["type"] != temp_filter:
                continue
            filtered_data.append(reg)

        if limit_filter != "All":
            limit = int(limit_filter)
            filtered_data = filtered_data[:limit]

        for reg in filtered_data:
            line = f"{reg['test_no']} | {reg['inflator_no']} | {reg['temperature_c']}°C | {reg['type']} | {reg['version']} | {reg['order']} | {reg['test_date']}"
            if reg["pressures"]:
                line += " | Pressure data available"
            else:
                line += " | No pressure data"
            self.list_results.insert(tk.END, line)
        self.filtered_workplace_data = filtered_data
        self.update_workplace_counters(filtered_data)

    def remove_workplace_orders_selected(self):
        selected_orders = {
            (version, order)
            for (version, order), var in self.order_vars.items()
            if var.get()
        }
        if not selected_orders:
            messagebox.showwarning(
                "Warning", "No orders selected to remove from Workplace."
            )
            return

        before = len(self.workplace_data)
        self.workplace_data = [
            reg
            for reg in self.workplace_data
            if (reg["version"], reg["order"]) not in selected_orders
        ]
        after = len(self.workplace_data)
        self.update_workplace_display()
        self.update_workplace_counters()
        messagebox.showinfo(
            "Success", f"{before - after} records removed from Workplace."
        )

    def update_workplace_counters(self, data=None):
        if data is None:
            data = self.workplace_data
        rt = lt = ht = 0
        for item in data:
            temp = item.get("type")
            if temp == "RT":
                rt += 1
            elif temp == "LT":
                lt += 1
            elif temp == "HT":
                ht += 1
        total = rt + lt + ht
        self.label_rt.config(text=f"RT: {rt}")
        self.label_lt.config(text=f"LT: {lt}")
        self.label_ht.config(text=f"HT: {ht}")
        self.label_total.config(text=f"Total: {total}")

    def clear_workplace(self):
        self.workplace_data.clear()
        self.list_results.delete(0, tk.END)
        messagebox.showinfo("Success", "Workplace cleared successfully.")
        self.update_workplace_counters()


if __name__ == "__main__":
    ExcelToJsonConverter()
