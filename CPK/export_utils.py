import json
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from tkinter import messagebox
import os


def adjust_column_widths(ws):
    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                col_letter = cell.column_letter
                current_width = column_widths.get(col_letter, 0)
                cell_len = len(str(cell.value))
                column_widths[col_letter] = max(current_width, cell_len)
    for col_letter, width in column_widths.items():
        adjusted_width = min(width * 1.2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width


def export_to_excel(data_by_temp, table_data, ms_points_dict):
    try:
        if not data_by_temp or not table_data or not ms_points_dict:
            raise ValueError("Invalid input data for Excel export")

        wb = Workbook()
        wb.remove(wb.active)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        bold_font = Font(bold=True)
        max_fill = PatternFill(
            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
        )
        mean_fill = PatternFill(
            start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
        )
        min_fill = PatternFill(
            start_color="CCE6FF", end_color="CCE6FF", fill_type="solid"
        )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{timestamp}.xlsx"

        for idx, temp in enumerate(["RT", "LT", "HT"]):
            if temp not in data_by_temp:
                continue
            records = data_by_temp[temp]
            ws = wb.create_sheet(title=temp)
            versions = set(r["version"] for r in records)
            version = ", ".join(versions) if len(versions) > 1 else list(versions)[0]
            total_inflators = len(records)

            ws.append(["Temperature", temp])
            ws.append(["Version", version])
            ws.append(["Total Inflators", total_inflators])
            ws.append([])

            for row in ws["A1:B3"]:
                for cell in row:
                    cell.alignment = center_alignment
                    cell.font = bold_font
                    cell.border = thin_border

            ms_points_str = ms_points_dict.get(temp, [])
            ws.append([""] + ms_points_str)

            header_row = 5
            for col_idx, ms in enumerate([""] + ms_points_str, 1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.value = ms
                cell.alignment = center_alignment
                cell.font = bold_font
                cell.border = thin_border

            for row_idx, (row_label, row_data) in enumerate(
                table_data[idx], start=header_row + 1
            ):
                ws.append([row_label] + row_data)
                for col_idx in range(1, len(row_data) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    if row_label == "Maximum":
                        cell.fill = max_fill
                    elif row_label == "Mean":
                        cell.fill = mean_fill
                    elif row_label == "Minimum":
                        cell.fill = min_fill

            ws.append([])
            ws.append(["Inflator Data"])
            ws.append(["Inflator No"] + ms_points_str)

            inflator_header_row = ws.max_row
            for col_idx, ms in enumerate(["Inflator No"] + ms_points_str, 1):
                cell = ws.cell(row=inflator_header_row, column=col_idx)
                cell.value = ms
                cell.alignment = center_alignment
                cell.font = bold_font
                cell.border = thin_border

            for r in records:
                if r["pressures"]:
                    row = [str(r["inflator_no"])]
                    for ms in ms_points_str:
                        val = r["pressures"].get(ms, np.nan)
                        row.append(f"{val:.2f}" if not np.isnan(val) else "-")
                    ws.append(row)
                    for col_idx in range(1, len(row) + 1):
                        cell = ws.cell(row=ws.max_row, column=col_idx)
                        cell.alignment = center_alignment
                        cell.border = thin_border

            adjust_column_widths(ws)

        wb.save(filename)
        messagebox.showinfo("Success", f"Report exported to {filename}")
    except Exception as e:
        messagebox.showerror("Error", f"Error exporting to Excel: {str(e)}")


def export_to_pdf(data_by_temp, table_data, ms_points_dict, json_file):
    try:
        if (
            not data_by_temp
            or not table_data
            or not ms_points_dict
            or not os.path.exists(json_file)
        ):
            raise ValueError("Invalid input data or JSON file not found")

        num_temps = len(data_by_temp)
        if num_temps == 0:
            raise ValueError("No temperature data to export")

        # Extend table_data if needed
        if len(table_data) < num_temps:
            table_data.extend([[]] * (num_temps - len(table_data)))

        versions = set()
        total_inflators = 0
        temp_counts = {"RT": 0, "LT": 0, "HT": 0}
        for temp, records in data_by_temp.items():
            if not records:
                continue
            versions.update(r["version"] for r in records)
            total_inflators += len(records)
            temp_counts[temp] = len(records)
        version_str = ", ".join(versions) if len(versions) > 1 else versions.pop()
        report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{timestamp}.pdf"

        # Ensure output directory is writable
        output_dir = os.path.dirname(json_file) or os.getcwd()
        try:
            with open(os.path.join(output_dir, "test_write.txt"), "w") as f:
                f.write("test")
            os.remove(os.path.join(output_dir, "test_write.txt"))
        except PermissionError:
            output_dir = os.path.expanduser("~/")
            filename = os.path.join(output_dir, f"report_{timestamp}.pdf")

        fig = plt.figure(figsize=(8.27, 11.69))  # A4 size
        fig.patch.set_facecolor("#fafafa")
        fig.subplots_adjust(left=0.15, right=0.85, top=0.92, bottom=0.08, hspace=0.5)

        ax_title = fig.add_axes([0.15, 0.92, 0.7, 0.06])
        ax_title.axis("off")
        ax_title.text(
            0.5,
            0.9,
            "Ballistic Tests Report",
            fontsize=12,
            ha="center",
            weight="bold",
        )
        ax_title.text(
            0.5,
            0.7,
            f"Version(s): {version_str} | Total Inflators: {total_inflators}",
            fontsize=6,
            ha="center",
        )
        ax_title.text(
            0.5,
            0.5,
            f"Temperatures: RT={temp_counts['RT']}, LT={temp_counts['LT']}, HT={temp_counts['HT']}",
            fontsize=6,
            ha="center",
        )

        axes = fig.subplots(
            num_temps * 2,
            1,
            gridspec_kw={"height_ratios": [2.5, 1] * num_temps},
        )
        if num_temps == 1:
            axes = [axes] if not isinstance(axes, np.ndarray) else axes

        for idx, temp in enumerate(["RT", "LT", "HT"]):
            if temp not in data_by_temp:
                continue
            records = data_by_temp[temp]
            ax_graph = axes[idx * 2]
            ax_table = axes[idx * 2 + 1]

            versions = set(r["version"] for r in records)
            version = ", ".join(versions) if len(versions) > 1 else list(versions)[0]
            total_inflators = len(records)

            ms_points = ms_points_dict.get(temp, [])
            ms_points_str = [str(ms) for ms in ms_points]
            if not ms_points_str:
                ax_table.axis("off")
                ax_graph.axis("off")
                continue

            pressure_matrix = []
            for r in records:
                p = []
                if r["pressures"]:
                    for ms in ms_points_str:
                        val = r["pressures"].get(ms, np.nan)
                        p.append(val)
                else:
                    p = [np.nan] * len(ms_points)
                pressure_matrix.append(p)
            pressure_matrix = np.array(pressure_matrix, dtype=np.float64)

            limits_max = []
            limits_min = []
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    data_json = json.load(f)
                sample_order = records[0]["order"]
                limits = (
                    data_json[version][sample_order]["temperatures"]
                    .get(temp, {})
                    .get("limits", {})
                )
                max_dict = limits.get("maximums", {})
                min_dict = limits.get("minimums", {})
                limits_max = [max_dict.get(str(ms), np.nan) for ms in ms_points]
                limits_min = [min_dict.get(str(ms), np.nan) for ms in ms_points]
            except (FileNotFoundError, json.JSONDecodeError, KeyError):
                limits_max = [np.nan] * len(ms_points)
                limits_min = [np.nan] * len(ms_points)

            mean = np.nanmean(pressure_matrix, axis=0)

            ax_graph.set_facecolor("#fafafa")
            for p in pressure_matrix:
                ax_graph.plot(
                    ms_points,
                    p,
                    color="#444444",
                    linewidth=1,
                    alpha=0.5,
                )
            ax_graph.plot(
                ms_points,
                limits_max,
                color="#d62728",
                linewidth=1.5,
                label="Maximum Limit",
                linestyle="--",
            )
            ax_graph.plot(
                ms_points,
                limits_min,
                color="#1f77b4",
                linewidth=1.5,
                label="Minimum Limit",
                linestyle="--",
            )
            ax_graph.plot(
                ms_points,
                mean,
                color="#2ca02c",
                linewidth=2,
                label="Mean",
                linestyle="-",
            )
            ax_graph.set_title(
                f"{temp} | Version: {version} | Inflators: {total_inflators}",
                fontsize=7,
                pad=5,
            )
            ax_graph.set_xlabel("Time (ms)", fontsize=5)
            ax_graph.set_ylabel("Pressure (bar)", fontsize=5)
            ax_graph.legend(loc="lower right", fontsize=5)
            ax_graph.grid(True, color="#cccccc", linestyle="--", linewidth=0.5)
            ax_graph.minorticks_on()
            ax_graph.grid(
                True,
                which="minor",
                color="#e0e0e0",
                linestyle=":",
                linewidth=0.3,
            )
            ax_graph.tick_params(axis="both", which="major", labelsize=5)

            def format_row(row):
                return [f"{v:.2f}" if not np.isnan(v) else "-" for v in row]

            table_data_rows = [
                format_row(limits_max),
                format_row(mean),
                format_row(limits_min),
            ]
            row_labels = ["Maximum", "Mean", "Minimum"]
            col_labels = ms_points_str

            try:
                if idx < len(table_data) and table_data[idx]:
                    table_data_rows = [row[1] for row in table_data[idx]]
                    row_labels = [row[0] for row in table_data[idx]]
            except IndexError:
                pass

            max_cols = min(len(col_labels), 12)
            col_labels = col_labels[:max_cols]
            table_data_rows = [row[:max_cols] for row in table_data_rows]

            cell_colors = [
                ["#ffcccc"] * max_cols,
                ["#ccffcc"] * max_cols,
                ["#cce6ff"] * max_cols,
            ]

            table = ax_table.table(
                cellText=table_data_rows,
                rowLabels=row_labels,
                colLabels=col_labels,
                cellColours=cell_colors,
                cellLoc="center",
                loc="center",
                bbox=[0.05, 0, 0.95, 1],
            )
            table.auto_set_font_size(False)
            table.set_fontsize(5)
            table.scale(1, 1.1)
            ax_table.axis("off")

        fig.text(
            0.15,
            0.03,
            f"Generated: {report_date}",
            fontsize=7,
            ha="left",
        )

        with PdfPages(filename) as pdf:
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)

        messagebox.showinfo("Success", f"Report exported to {filename}")
    except Exception as e:
        messagebox.showerror("Error", f"Error exporting to PDF: {str(e)}")
