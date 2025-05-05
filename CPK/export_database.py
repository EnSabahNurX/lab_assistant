import os
import json
from datetime import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from tkinter import messagebox
import traceback
from openpyxl.worksheet.protection import SheetProtection
import config


def export_database_to_excel(self):
    try:
        if not os.path.exists(self.json_file):
            messagebox.showerror("Error", config.ERROR_MESSAGES["no_database_found"])
            return

        with open(self.json_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        if not data:
            messagebox.showwarning("Warning", config.ERROR_MESSAGES["empty_database"])
            return

        wb = Workbook()
        wb.remove(wb.active)

        # Aba de Resumo
        ws_summary = wb.create_sheet(title="Summary", index=0)
        ws_summary.append(["Ballistic Tests Database Summary"])
        ws_summary.append(
            [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
        )
        ws_summary.append([])

        summary_stats = {
            "total_orders": 0,
            "total_tests": 0,
            "tests_by_temp": {temp: 0 for temp in config.TEMPERATURE_TYPES},
        }
        for version in data:
            summary_stats["total_orders"] += len(data[version])
            for order, order_data in data[version].items():
                for temp_type, temp_data in order_data["temperatures"].items():
                    summary_stats["tests_by_temp"][temp_type] += len(
                        temp_data.get("tests", [])
                    )
                    summary_stats["total_tests"] += len(temp_data.get("tests", []))

        ws_summary.append(["Total Orders", summary_stats["total_orders"]])
        ws_summary.append(["Total Tests", summary_stats["total_tests"]])
        for temp in config.TEMPERATURE_TYPES:
            ws_summary.append([f"{temp} Tests", summary_stats["tests_by_temp"][temp]])
        ws_summary.append([])
        ws_summary.append(
            ["Note: Detailed data is available in version-specific sheets."]
        )

        for row in ws_summary["A1:A2"]:
            for cell in row:
                cell.font = config.TITLE_FONT
                cell.fill = config.TITLE_FILL
                cell.alignment = config.CENTER_ALIGNMENT
                cell.border = config.THICK_BORDER
        for row in ws_summary["A4:B8"]:
            for cell in row:
                cell.font = config.METADATA_FONT
                cell.fill = config.METADATA_FILL
                cell.alignment = config.CENTER_ALIGNMENT
                cell.border = config.THIN_BORDER
        ws_summary["A10"].font = config.NOTE_FONT
        ws_summary["A10"].alignment = config.CENTER_ALIGNMENT

        # Ajusta largura e altura da aba de resumo
        column_widths = {}
        for row in ws_summary.iter_rows():
            for cell in row:
                if cell.value:
                    col_letter = cell.column_letter
                    current_width = column_widths.get(col_letter, 0)
                    cell_len = len(str(cell.value)) + 2
                    column_widths[col_letter] = max(current_width, cell_len)
        for col_letter, width in column_widths.items():
            adjusted_width = min(max(width * 1.2, 10), 30)
            ws_summary.column_dimensions[col_letter].width = adjusted_width

        for row in ws_summary.iter_rows():
            max_height = 15
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    lines = cell.value.count("\n") + 1
                    width = column_widths.get(cell.column_letter, 10)
                    chars_per_line = max(len(line) for line in cell.value.split("\n"))
                    if chars_per_line > width:
                        lines += chars_per_line // width
                    max_height = max(max_height, lines * 15)
            ws_summary.row_dimensions[row[0].row].height = max_height

        for version in sorted(data.keys()):
            ws = wb.create_sheet(title=version)
            total_orders = len(data[version])
            total_tests = 0

            # Metadados (A1:D1)
            ws.append(
                [
                    "Ballistic Tests Database 🌐",
                    f"Version: {version}",
                    f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                    f"Total Orders: {total_orders}",
                ]
            )
            for cell in ws["A1:D1"][0]:
                cell.font = config.TITLE_FONT
                cell.fill = config.TITLE_FILL
                cell.alignment = config.CENTER_ALIGNMENT
                cell.border = config.THICK_BORDER

            # Cabeçalhos principais (row 2)
            main_headers = [
                "Order Number 📋",
                "Production Order 🏭",
                "Propellant Lot Number 🧪",
                "Test Date 📅",
                "Test Number 🔢",
                "Inflator Number ⚙️",
            ]
            ws.append(main_headers)
            header_row = 2
            for col_idx, header in enumerate(main_headers, 1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.font = config.HEADER_FONT
                cell.fill = config.METADATA_FILL
                cell.border = config.THIN_BORDER
                cell.alignment = config.CENTER_ALIGNMENT

            # Coleta de pontos de tempo (ms) e limites por temperatura
            ms_points_by_temp = {temp: set() for temp in config.TEMPERATURE_TYPES}
            limits_by_temp = {
                temp: {"max": {}, "min": {}} for temp in config.TEMPERATURE_TYPES
            }
            for order, order_data in data[version].items():
                for temp_type, temp_data in order_data["temperatures"].items():
                    for pressure_entry in temp_data.get("pressure_data", []):
                        ms_points_by_temp[temp_type].update(
                            str(key) for key in pressure_entry["pressures"].keys()
                        )
                    limits = temp_data.get("limits", {})
                    limits_by_temp[temp_type]["max"].update(limits.get("maximums", {}))
                    limits_by_temp[temp_type]["min"].update(limits.get("minimums", {}))

            for temp in ms_points_by_temp:
                ms_points_by_temp[temp] = sorted(ms_points_by_temp[temp], key=int)

            # Cabeçalhos de temperatura e tempos (rows 3-4)
            temp_headers_row = header_row + 1
            col_idx = len(main_headers) + 1
            for temp in config.TEMPERATURE_TYPES:
                if not ms_points_by_temp[temp]:
                    continue
                ms_points = ms_points_by_temp[temp]
                ws.cell(row=temp_headers_row, column=col_idx).value = f"{temp} Data 🌡️"
                ws.merge_cells(
                    start_row=temp_headers_row,
                    start_column=col_idx,
                    end_row=temp_headers_row,
                    end_column=col_idx + len(ms_points) - 1,
                )
                cell = ws.cell(row=temp_headers_row, column=col_idx)
                cell.font = config.HEADER_FONT
                cell.alignment = config.CENTER_ALIGNMENT
                cell.border = config.THIN_BORDER
                cell.fill = (
                    config.RT_FILL
                    if temp == "RT"
                    else config.LT_FILL if temp == "LT" else config.HT_FILL
                )

                for ms in ms_points:
                    cell = ws.cell(row=temp_headers_row + 1, column=col_idx)
                    cell.value = f"{ms} ms ⏱️"
                    cell.font = config.HEADER_FONT
                    cell.alignment = config.CENTER_ALIGNMENT
                    cell.border = config.THIN_BORDER
                    cell.fill = (
                        config.RT_FILL
                        if temp == "RT"
                        else config.LT_FILL if temp == "LT" else config.HT_FILL
                    )
                    col_idx += 1

            # Dados (row 5+)
            row_idx = temp_headers_row + 2
            start_merge_row = row_idx
            orders_sorted = sorted(
                data[version].items(),
                key=lambda x: (
                    datetime.strptime(
                        x[1].get("metadata", {}).get("test_date", "0000-00-00"),
                        "%Y-%m-%d",
                    )
                    if x[1].get("metadata", {}).get("test_date", "0000-00-00")
                    != "0000-00-00"
                    else datetime.min
                ),
                reverse=True,
            )

            for order, order_data in orders_sorted:
                metadata = order_data.get("metadata", {})
                production_order = metadata.get("production_order", "N/A")
                propellant_lot = metadata.get("propellant_lot_number", "N/A")
                test_date = metadata.get("test_date", "N/A")
                temp_types = config.TEMPERATURE_TYPES
                tests_by_temp = {}
                pressure_maps = {}

                for temp_type in temp_types:
                    if temp_type not in order_data["temperatures"]:
                        continue
                    temp_data = order_data["temperatures"][temp_type]
                    temp_c = temp_data.get("temperature_c", "N/A")
                    tests = temp_data.get("tests", [])
                    pressure_data = temp_data.get("pressure_data", [])
                    pressure_maps[temp_type] = {
                        entry["inflator_no"]: entry["pressures"]
                        for entry in pressure_data
                    }
                    tests_by_temp[temp_type] = [
                        {
                            "test_no": test.get("test_no", "N/A"),
                            "inflator_no": test.get("inflator_no", "N/A"),
                            "temp_c": str(temp_c),
                        }
                        for test in tests
                    ]

                total_rows_for_order = sum(
                    len(tests_by_temp.get(temp, [])) for temp in temp_types
                )
                if total_rows_for_order == 0:
                    continue

                for temp_type in temp_types:
                    if temp_type not in tests_by_temp:
                        continue
                    for test in tests_by_temp[temp_type]:
                        row = [
                            str(order),
                            str(production_order),
                            str(propellant_lot),
                            str(test_date),
                            str(test["test_no"]),
                            str(test["inflator_no"]),
                        ]
                        for temp in config.TEMPERATURE_TYPES:
                            ms_points = ms_points_by_temp[temp]
                            pressures = pressure_maps.get(temp, {}).get(
                                test["inflator_no"], {}
                            )
                            if temp == temp_type:
                                for ms in ms_points:
                                    val = pressures.get(str(ms), np.nan)
                                    formatted_val = (
                                        f"{val:.2f}" if not np.isnan(val) else "-"
                                    )
                                    row.append(formatted_val)
                            else:
                                row.extend([""] * len(ms_points))
                        ws.append(row)
                        total_tests += 1

                        for col_idx in range(1, len(row) + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.font = config.DATA_FONT
                            cell.border = config.THIN_BORDER
                            cell.alignment = config.CENTER_ALIGNMENT
                            if row_idx % 2 == 0:
                                cell.fill = config.ALT_FILL
                            if col_idx > len(main_headers) and cell.value not in [
                                "",
                                "-",
                            ]:
                                try:
                                    val = float(cell.value)
                                    ms_idx = col_idx - len(main_headers) - 1
                                    temp_idx = next(
                                        i
                                        for i, t in enumerate(config.TEMPERATURE_TYPES)
                                        if ms_idx
                                        < sum(
                                            len(ms_points_by_temp[t])
                                            for t in config.TEMPERATURE_TYPES[: i + 1]
                                        )
                                    )
                                    temp = config.TEMPERATURE_TYPES[temp_idx]
                                    ms = ms_points_by_temp[temp][
                                        ms_idx
                                        - sum(
                                            len(ms_points_by_temp[t])
                                            for t in config.TEMPERATURE_TYPES[:temp_idx]
                                        )
                                    ]
                                    max_limit = limits_by_temp[temp]["max"].get(str(ms))
                                    min_limit = limits_by_temp[temp]["min"].get(str(ms))
                                    if max_limit is not None and val > float(max_limit):
                                        cell.fill = config.WARNING_FILL
                                    elif min_limit is not None and val < float(
                                        min_limit
                                    ):
                                        cell.fill = config.LOW_FILL
                                except (ValueError, IndexError):
                                    pass
                        row_idx += 1

                if total_rows_for_order > 1:
                    for col in range(1, 5):
                        ws.merge_cells(
                            start_row=start_merge_row,
                            start_column=col,
                            end_row=start_merge_row + total_rows_for_order - 1,
                            end_column=col,
                        )
                start_merge_row = row_idx

            # Total de testes
            ws.append([])
            ws.append([f"Total Tests: {total_tests}"])
            for cell in ws[f"A{row_idx + 1}:A{row_idx + 1}"]:
                cell[0].font = config.METADATA_FONT
                cell[0].fill = config.METADATA_FILL
                cell[0].alignment = config.CENTER_ALIGNMENT
                cell[0].border = config.THICK_BORDER

            # Notas de rodapé
            ws.append([])
            ws.append(["Notes:"])
            ws.append(["- Values in red indicate pressures above the maximum limit."])
            ws.append(["- Values in blue indicate pressures below the minimum limit."])
            ws.append(["- Use filters to sort or analyze data."])
            for row in ws[f"A{row_idx + 3}:A{row_idx + 6}"]:
                for cell in row:
                    cell.font = config.NOTE_FONT
                    cell.alignment = config.CENTER_ALIGNMENT
                    cell.border = config.THIN_BORDER

            # Ajusta largura das colunas
            column_widths = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        col_letter = cell.column_letter
                        current_width = column_widths.get(col_letter, 0)
                        cell_len = len(str(cell.value)) + 2
                        column_widths[col_letter] = max(current_width, cell_len)
            for col_letter, width in column_widths.items():
                adjusted_width = min(max(width * 1.2, 10), 30)
                ws.column_dimensions[col_letter].width = adjusted_width

            # Ajusta altura das linhas
            for row in ws.iter_rows():
                row_number = row[0].row
                max_height = 15
                height_scale = 15
                if row_number == 1:
                    height_scale = 18
                elif row_number in [2, 3]:
                    height_scale = 15
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        lines = cell.value.count("\n") + 1
                        width = column_widths.get(cell.column_letter, 10)
                        chars_per_line = max(
                            len(line) for line in cell.value.split("\n")
                        )
                        if chars_per_line > width:
                            lines += chars_per_line // width
                        max_height = max(max_height, lines * height_scale)
                ws.row_dimensions[row_number].height = max_height

            # Fixa cabeçalhos
            ws.freeze_panes = "G5"

            # Adiciona filtros
            table_range = f"A{header_row}:{get_column_letter(ws.max_column)}{row_idx}"
            tab = Table(displayName=f"Table_{version}", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)

            # Protege a planilha
            ws.protection = SheetProtection(
                sheet=True,
                formatCells=False,
                formatColumns=False,
                formatRows=False,
                insertRows=False,
                insertColumns=False,
                deleteRows=False,
                deleteColumns=False,
                sort=True,
                autoFilter=True,
            )

        # Aba de instruções
        ws_instructions = wb.create_sheet(title="Instructions", index=1)
        ws_instructions.append(
            ["Instructions for Using the Ballistic Tests Database 📖"]
        )
        ws_instructions.append([])
        instructions = [
            "Welcome to the Ballistic Tests Database Export!",
            "This Excel file contains test data organized by version, with each version in a separate sheet.",
            "",
            "Key Features:",
            "- The 'Summary' sheet provides an overview of total orders and tests.",
            "- Each version sheet (e.g., V1, V2) contains detailed test data.",
            "- Data is grouped by Order Number, sorted by test date (most recent first).",
            "- Within each order, tests are ordered by temperature: RT (Room Temperature), LT (Low Temperature), HT (High Temperature).",
            "- Pressure values are listed in milliseconds (ms) under each temperature.",
            "- Headers are frozen for easy navigation (scroll while keeping headers visible).",
            "- Filters are enabled for sorting and analyzing data.",
            "",
            "Color Coding:",
            "- RT Headers: Light Green 🌡️",
            "- LT Headers: Light Blue 🌡️",
            "- HT Headers: Light Orange 🌡️",
            "- Red cells indicate pressures above the maximum limit 🚨",
            "- Blue cells indicate pressures below the minimum limit ❄️",
            "- Alternating rows have a light gray background for readability.",
            "",
            "Tips:",
            "- Use the 'Summary' sheet to get a quick overview.",
            "- Apply filters to explore specific orders or tests.",
            "- Protected sheets prevent accidental edits but allow filtering and sorting.",
            "- Contact the Inflator Lab team for support if needed.",
        ]
        for line in instructions:
            ws_instructions.append([line])
        for row in ws_instructions["A1:A1"]:
            for cell in row:
                cell.font = config.TITLE_FONT
                cell.fill = config.TITLE_FILL
                cell.alignment = config.CENTER_ALIGNMENT
                cell.border = config.THICK_BORDER
        for row in ws_instructions["A2:A{}".format(ws_instructions.max_row)]:
            for cell in row:
                cell.font = config.DATA_FONT
                cell.alignment = config.CENTER_ALIGNMENT
                cell.border = config.THIN_BORDER

        # Ajusta largura e altura da aba de instruções
        column_widths = {}
        for row in ws_instructions.iter_rows():
            for cell in row:
                if cell.value:
                    col_letter = cell.column_letter
                    current_width = column_widths.get(col_letter, 0)
                    cell_len = len(str(cell.value)) + 2
                    column_widths[col_letter] = max(current_width, cell_len)
        for col_letter, width in column_widths.items():
            adjusted_width = min(max(width * 1.2, 10), 100)
            ws_instructions.column_dimensions[col_letter].width = adjusted_width

        for row in ws_instructions.iter_rows():
            max_height = 15
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    lines = cell.value.count("\n") + 1
                    width = column_widths.get(cell.column_letter, 10)
                    chars_per_line = max(len(line) for line in cell.value.split("\n"))
                    if chars_per_line > width:
                        lines += chars_per_line // width
                    max_height = max(max_height, lines * 15)
            ws_instructions.row_dimensions[row[0].row].height = max_height

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Data_{timestamp}.xlsx"
        wb.save(filename)
        messagebox.showinfo("Success", f"Database exported to {filename}")

    except Exception as e:
        error_message = config.ERROR_MESSAGES["export_error"].format(error=str(e))
        print(error_message)
        messagebox.showerror("Error", error_message)
