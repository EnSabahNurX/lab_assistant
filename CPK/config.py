# config.py
"""
Configuration constants and settings for the Ballistic Tests Database application.
Centralizes file paths, Excel processing parameters, UI settings, and other key values
to facilitate maintenance and updates.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# File Paths and Database Settings
JSON_FILE = "Data.json"
BACKUP_FOLDER = "Backup"
EXCEL_FOLDER = r"H:\TEAMS\Inflator_Lab\0_Evaluations\vi"

# Excel Processing Constants (for process_grafik and process_datenblatt)
MIN_LIMIT_ROW = 51
MAX_LIMIT_ROW = 55
PRESSURE_DATA_START_ROW = 60
MIN_COLUMN = 3
MAX_COLUMN = 151
EXPECTED_SHEET_NAME = "Grafik"  # Expected sheet name for grafik files
DATENBLATT_SHEET_NAME = "Datenblatt"  # Expected sheet name for datenblatt files

# Temperature Types
TEMPERATURE_TYPES = ["RT", "LT", "HT"]

# Backup Settings
BACKUP_PREFIX = "Backup_Data_"
BACKUP_EXTENSION = ".json"

# Date Filter Periods (for main.py)
DATE_FILTER_PERIODS = {
    "All": None,
    "Last 30 days": 30,
    "Last 60 days": 60,
    "Last 90 days": 90,
}

# Excel Styling Constants (for export_database.py)
TITLE_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
METADATA_FONT = Font(name="Calibri", size=10, bold=True)
HEADER_FONT = Font(name="Calibri", size=10, bold=True)
DATA_FONT = Font(name="Calibri", size=10)
NOTE_FONT = Font(name="Calibri", size=9, italic=True)

TITLE_FILL = PatternFill(
    start_color="4682B4", end_color="4682B4", fill_type="solid"
)  # Steel blue
METADATA_FILL = PatternFill(
    start_color="E6F0FA", end_color="E6F0FA", fill_type="solid"
)  # Light blue
RT_FILL = PatternFill(
    start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
)  # Light green
LT_FILL = PatternFill(
    start_color="CCE6FF", end_color="CCE6FF", fill_type="solid"
)  # Light blue
HT_FILL = PatternFill(
    start_color="FFE6CC", end_color="FFE6CC", fill_type="solid"
)  # Light orange
ALT_FILL = PatternFill(
    start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
)  # Light gray
WARNING_FILL = PatternFill(
    start_color="FF9999", end_color="FF9999", fill_type="solid"
)  # Red
LOW_FILL = PatternFill(
    start_color="99CCFF", end_color="99CCFF", fill_type="solid"
)  # Blue

CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
THICK_BORDER = Border(
    left=Side(style="thick"),
    right=Side(style="thick"),
    top=Side(style="thick"),
    bottom=Side(style="thick"),
)

# Error Messages
ERROR_MESSAGES = {
    "no_database_found": "Database file not found.",
    "empty_database": "Database is empty.",
    "export_error": "Error exporting database: {error}",
    "invalid_date_format": "Invalid date format. Use YYYY-MM-DD.",
    "invalid_date_range": "End date must be after start date.",
}
